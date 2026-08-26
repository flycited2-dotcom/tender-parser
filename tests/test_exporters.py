import json
from dataclasses import replace
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from tender_parser.exporters.excel import (
    export_excel,
    load_customer_rows,
    load_manual_selections,
    sort_for_review,
)
from tender_parser.exporters.html_report import export_html_report
from tender_parser.exporters.json_exporter import export_json, export_run_report
from tender_parser.models import TenderRecord
from tender_parser.run_report import SourceFetchResult, SourceHealth


def make_tender(status: str) -> TenderRecord:
    priority = {"matched": "hot", "review": "review", "excluded": "excluded"}[status]
    return TenderRecord(
        title="Поставка МФУ",
        url="https://example.test/tender-1/",
        source="test",
        tender_number="1",
        customer="Заказчик",
        region="Республика Крым",
        price=45_000.0,
        deadline=datetime(2026, 5, 25, 10, 0),
        filter_status=status,
        review_priority=priority,
        category="Компьютерная техника и периферия" if status != "excluded" else None,
        include_reason="ok" if status != "excluded" else "",
        exclude_reason="" if status == "matched" else "регион не найден",
        match_confidence="точное" if status == "matched" else "ручная проверка",
        discovered_at=datetime(2026, 5, 19, 12, 0),
    )


NOW = datetime(2026, 5, 19, 12, 0)


def test_export_excel_creates_expected_sheets(tmp_path: Path) -> None:
    output = tmp_path / "tenders.xlsx"
    export_excel(
        [make_tender("matched")],
        [make_tender("review")],
        [replace(make_tender("review"), review_priority="wide")],
        [make_tender("excluded")],
        output,
        now=NOW,
    )

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Дашборд", "Горячие", "На проверку", "Широкий хвост", "Отсеянные"]
    hot_sheet = workbook["Горячие"]
    assert hot_sheet["A1"].value == "приоритет"
    assert hot_sheet["B1"].value == "новый"
    assert hot_sheet["A2"].value == "Горячий"
    assert hot_sheet["D2"].value == "Поставка МФУ"
    assert workbook["На проверку"]["D2"].value == "Поставка МФУ"
    assert workbook["Широкий хвост"]["D2"].value == "Поставка МФУ"


def test_export_excel_adds_regional_and_customer_sheets(tmp_path: Path) -> None:
    output = tmp_path / "tenders_2026-05-19.xlsx"
    customer_row = [
        "org-key",
        "ГБУ РК Тест",
        "Государственная / муниципальная",
        "Республика Крым",
        "9100000000",
        "",
        "",
        "mail@example.test",
        "+7 978 000-00-00",
        "",
        "https://customer.example.test",
        "https://zakupki.gov.ru/",
        "https://example.test/tender-1/",
        "19.05.2026",
        "Проверен",
        "",
    ]
    regional = replace(
        make_tender("excluded"),
        title="Поставка лекарственных препаратов",
        exclude_reason="стоп-тема: лекарственные препараты",
    )

    export_excel(
        [make_tender("matched")],
        [],
        [],
        [regional],
        output,
        now=NOW,
        regional_tenders=[make_tender("matched"), regional],
        customer_rows=[customer_row],
    )

    workbook = load_workbook(output)
    assert workbook["Все региональные"].max_row == 3
    assert workbook["Все региональные"]["D3"].value == "Поставка лекарственных препаратов"
    customers = workbook["Потенциальные заказчики"]
    assert customers["H2"].value == "mail@example.test"
    assert customers["M2"].hyperlink.target == "https://example.test/tender-1/"
    assert load_customer_rows(tmp_path)[0][1] == "ГБУ РК Тест"


def test_export_excel_data_sheet_is_filterable_and_styled(tmp_path: Path) -> None:
    output = tmp_path / "tenders.xlsx"
    export_excel([make_tender("matched")], [], [], [], output, now=NOW)

    sheet = load_workbook(output)["Горячие"]
    assert sheet.auto_filter.ref is not None
    assert sheet.freeze_panes == "E2"
    assert sheet["D2"].hyperlink.target == "https://example.test/tender-1/"
    assert sheet["F2"].value == 45_000.0
    assert "₽" in sheet["F2"].number_format
    assert sheet["G2"].value == datetime(2026, 5, 25, 10, 0)
    assert sheet["A2"].fill.start_color.rgb == "FFFBE0D6"
    assert sheet["C2"].value == 6
    assert sheet["C2"].fill.start_color.rgb == "FFFEECCA"
    assert len(sheet.data_validations.dataValidation) == 1


def test_export_excel_builds_dashboard_with_kpi_and_sources(tmp_path: Path) -> None:
    output = tmp_path / "tenders.xlsx"
    health = [SourceHealth(source="rts-cabinet", status="ok", found=1, elapsed_seconds=0.5)]
    export_excel(
        [make_tender("matched")],
        [make_tender("review")],
        [],
        [make_tender("excluded")],
        output,
        new_tenders=[make_tender("matched")],
        now=NOW,
        source_health=health,
    )

    dashboard = load_workbook(output)["Дашборд"]
    labels = {dashboard.cell(row=4, column=index).value for index in range(1, 7)}
    assert {"Горячие", "На проверку", "Новые для CRM"} <= labels
    values = [dashboard.cell(row=5, column=index).value for index in range(1, 7)]
    assert 1 in values
    column_a = [cell.value for cell in dashboard["A"] if cell.value]
    assert "Actionable по источникам" in column_a
    assert "Топ горячих" in column_a
    assert "Здоровье источников" in column_a
    assert "test" in column_a
    assert "rts-cabinet" in column_a


def test_data_sheets_have_grid_borders_and_computed_row_heights(tmp_path: Path) -> None:
    output = tmp_path / "tenders.xlsx"
    long_title = "Поставка МФУ, принтеров, сканеров и расходных материалов " * 3
    export_excel([replace(make_tender("matched"), title=long_title)], [], [], [], output, now=NOW)

    sheet = load_workbook(output)["Горячие"]
    assert sheet["A1"].border.bottom.style == "thin"
    assert sheet["D2"].border.bottom.style == "thin"
    assert sheet.row_dimensions[2].height is not None
    assert sheet.row_dimensions[2].height > 20


def test_dashboard_tables_have_borders_and_wide_title_column(tmp_path: Path) -> None:
    output = tmp_path / "tenders.xlsx"
    export_excel([make_tender("matched")], [], [], [make_tender("excluded")], output, now=NOW)

    dashboard = load_workbook(output)["Дашборд"]
    assert dashboard.column_dimensions["A"].width >= 50
    header_row = next(
        cell.row for cell in dashboard["A"] if cell.value == "Источник"
    )
    assert dashboard.cell(row=header_row, column=1).border.bottom.style == "thin"
    assert dashboard.cell(row=header_row + 1, column=1).border.bottom.style == "thin"


def test_export_excel_strips_illegal_characters(tmp_path: Path) -> None:
    output = tmp_path / "tenders.xlsx"
    dirty = replace(
        make_tender("matched"),
        title="Поставка\x0bМФУ\x01",
        include_reason="ok\x00",
    )

    export_excel([dirty], [], [], [], output, now=NOW)

    sheet = load_workbook(output)["Горячие"]
    assert sheet["D2"].value == "ПоставкаМФУ"


def test_manual_selections_read_by_mtime_across_fallback_files(tmp_path: Path) -> None:
    import os

    old_fallback = tmp_path / "tenders_2026-07-03_090000.xlsx"
    export_excel([make_tender("matched")], [], [], [], old_fallback, now=NOW)
    os.utime(old_fallback, (1_000_000, 1_000_000))

    main_file = tmp_path / "tenders_2026-07-03.xlsx"
    export_excel([make_tender("matched")], [], [], [], main_file, now=NOW)
    workbook = load_workbook(main_file)
    sheet = workbook["Горячие"]
    headers = [cell.value for cell in sheet[1]]
    sheet.cell(row=2, column=headers.index("мой_выбор") + 1, value="✅ Беру")
    workbook.save(main_file)

    selections = load_manual_selections(tmp_path)

    assert selections[("test", "1")] == ("✅ Беру", "")


def test_manual_selections_prefer_data_sheets_over_my_pick(tmp_path: Path) -> None:
    path = tmp_path / "tenders_2026-07-03.xlsx"
    export_excel(
        [make_tender("matched")],
        [],
        [],
        [],
        path,
        now=NOW,
        manual_selections={("test", "1"): ("✅ Беру", "")},
    )
    workbook = load_workbook(path)
    hot = workbook["Горячие"]
    headers = [cell.value for cell in hot[1]]
    hot.cell(row=2, column=headers.index("мой_выбор") + 1, value="❌ Мимо")
    workbook.save(path)  # «Мой отбор» сохранил старую ✅, пользователь правил «Горячие»

    selections = load_manual_selections(tmp_path)

    assert selections[("test", "1")][0] == "❌ Мимо"


def test_manual_selections_normalize_numeric_numbers(tmp_path: Path) -> None:
    path = tmp_path / "tenders_2026-07-03.xlsx"
    export_excel([make_tender("matched")], [], [], [], path, now=NOW)
    workbook = load_workbook(path)
    sheet = workbook["Горячие"]
    headers = [cell.value for cell in sheet[1]]
    sheet.cell(row=2, column=headers.index("номер") + 1, value=123.0)
    sheet.cell(row=2, column=headers.index("мой_выбор") + 1, value="✅ Беру")
    workbook.save(path)

    selections = load_manual_selections(tmp_path)

    assert ("test", "123") in selections


def test_html_report_neutralizes_javascript_urls(tmp_path: Path) -> None:
    output = tmp_path / "latest.html"
    evil = replace(make_tender("matched"), url="javascript:alert(document.cookie)")
    export_html_report(
        [evil],
        output,
        source_report=SourceFetchResult(),
        raw_count=1,
        unique_count=1,
        new_count=0,
    )

    html = output.read_text(encoding="utf-8")
    assert "javascript:" not in html


def test_export_json_sanitizes_non_finite_prices(tmp_path: Path) -> None:
    output = tmp_path / "latest.json"
    weird = replace(make_tender("matched"), price=float("nan"))

    export_json([weird], output)

    text = output.read_text(encoding="utf-8")
    assert "NaN" not in text
    assert json.loads(text)["items"][0]["price"] is None


def test_export_excel_falls_back_when_target_is_locked(tmp_path: Path) -> None:
    locked = tmp_path / "tenders_2026-07-03.xlsx"
    locked.mkdir()  # каталог с именем файла даёт PermissionError при записи, как открытый Excel

    result = export_excel([make_tender("matched")], [], [], [], locked, now=NOW)

    assert result != locked
    assert result.parent == tmp_path
    assert result.name.startswith("tenders_2026-07-03_")
    assert result.exists()


def test_new_tenders_are_marked_on_all_sheets(tmp_path: Path) -> None:
    output = tmp_path / "tenders.xlsx"
    tender = make_tender("matched")
    export_excel(
        [tender],
        [],
        [],
        [],
        output,
        new_tenders=[tender],
        now=NOW,
        new_keys={tender.unique_key},
    )

    workbook = load_workbook(output)
    hot = workbook["Горячие"]
    headers = [cell.value for cell in hot[1]]
    new_col = headers.index("новый") + 1
    assert hot.cell(row=2, column=new_col).value == "🆕"
    assert workbook["Новые"].cell(row=2, column=new_col).value == "🆕"


def test_old_tenders_have_no_new_mark(tmp_path: Path) -> None:
    output = tmp_path / "tenders.xlsx"
    export_excel([make_tender("matched")], [], [], [], output, now=NOW, new_keys=set())

    hot = load_workbook(output)["Горячие"]
    headers = [cell.value for cell in hot[1]]
    assert hot.cell(row=2, column=headers.index("новый") + 1).value in ("", None)


def test_manual_selections_survive_reexport(tmp_path: Path) -> None:
    first_path = tmp_path / "tenders_2026-07-03.xlsx"
    export_excel([make_tender("matched")], [], [], [], first_path, now=NOW)

    workbook = load_workbook(first_path)
    sheet = workbook["Горячие"]
    headers = [cell.value for cell in sheet[1]]
    choice_col = headers.index("мой_выбор") + 1
    comment_col = headers.index("комментарий") + 1
    sheet.cell(row=2, column=choice_col, value="✅ Беру")
    sheet.cell(row=2, column=comment_col, value="перезвонить заказчику")
    workbook.save(first_path)

    selections = load_manual_selections(tmp_path)
    assert selections[("test", "1")] == ("✅ Беру", "перезвонить заказчику")

    second_path = tmp_path / "tenders_2026-07-04.xlsx"
    export_excel(
        [make_tender("matched")], [], [], [], second_path, now=NOW, manual_selections=selections
    )

    second = load_workbook(second_path)
    hot = second["Горячие"]
    assert hot.cell(row=2, column=choice_col).value == "✅ Беру"
    assert hot.cell(row=2, column=comment_col).value == "перезвонить заказчику"
    assert "Мой отбор" in second.sheetnames
    picked = second["Мой отбор"]
    assert picked.cell(row=2, column=headers.index("название") + 1).value == "Поставка МФУ"


def test_load_manual_selections_handles_missing_dir(tmp_path: Path) -> None:
    assert load_manual_selections(tmp_path / "missing") == {}


def test_dashboard_lists_exclusion_reasons_and_closing_soon(tmp_path: Path) -> None:
    output = tmp_path / "tenders.xlsx"
    closing = replace(make_tender("matched"), deadline=datetime(2026, 5, 21, 10, 0))
    export_excel(
        [closing],
        [],
        [],
        [make_tender("excluded")],
        output,
        now=NOW,
    )

    dashboard = load_workbook(output)["Дашборд"]
    column_a = [cell.value for cell in dashboard["A"] if cell.value]
    assert "Топ причин отсева" in column_a
    assert any("регион не найден" in str(value) for value in column_a)
    assert "Скоро закрываются" in column_a


def test_export_excel_adds_new_sheet_when_new_tenders_are_given(tmp_path: Path) -> None:
    output = tmp_path / "tenders.xlsx"

    export_excel(
        [make_tender("matched")],
        [],
        [],
        [],
        output,
        new_tenders=[make_tender("matched")],
        now=NOW,
    )

    workbook = load_workbook(output)
    assert workbook.sheetnames[:2] == ["Дашборд", "Новые"]
    assert workbook["Новые"]["D2"].value == "Поставка МФУ"


def test_sort_for_review_orders_by_priority_deadline_price_and_discovery() -> None:
    hot_late = replace(
        make_tender("matched"),
        tender_number="2",
        deadline=datetime(2026, 5, 30),
        price=100_000.0,
    )
    hot_soon = replace(
        make_tender("matched"),
        tender_number="1",
        deadline=datetime(2026, 5, 20),
        price=40_000.0,
    )
    wide = replace(
        make_tender("review"),
        tender_number="3",
        review_priority="wide",
        deadline=datetime(2026, 5, 19),
        price=1_000_000.0,
    )

    result = sort_for_review([wide, hot_late, hot_soon])

    assert [item.tender_number for item in result] == ["1", "2", "3"]


def test_sort_for_review_prioritizes_crimea_before_new_target_regions() -> None:
    crimea = replace(
        make_tender("matched"),
        tender_number="crimea",
        region="Республика Крым",
        deadline=datetime(2026, 5, 30),
    )
    zaporizhzhia = replace(
        make_tender("matched"),
        tender_number="zaporizhzhia",
        region="Запорожская область",
        deadline=datetime(2026, 5, 25),
    )
    kherson = replace(
        make_tender("matched"),
        tender_number="kherson",
        region="Херсонская область",
        deadline=datetime(2026, 5, 20),
    )

    result = sort_for_review([kherson, zaporizhzhia, crimea])

    assert [item.tender_number for item in result] == [
        "crimea",
        "kherson",
        "zaporizhzhia",
    ]


def test_export_json_writes_matched_tenders(tmp_path: Path) -> None:
    output = tmp_path / "latest.json"
    export_json([make_tender("matched")], output)

    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["count"] == 1
    assert data["items"][0]["title"] == "Поставка МФУ"
    assert data["items"][0]["filter_status"] == "matched"
    assert data["items"][0]["match_confidence"] == "точное"
    assert data["items"][0]["review_priority"] == "hot"


def test_export_run_report_writes_source_health(tmp_path: Path) -> None:
    output = tmp_path / "run_report.json"
    report = SourceFetchResult(
        health=[
            SourceHealth(
                source="EisZakupkiSource",
                status="ok",
                found=12,
                elapsed_seconds=1.25,
            )
        ]
    )

    export_run_report(report, output, raw_count=12, unique_count=10, new_count=3)

    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["summary"] == {"raw_count": 12, "unique_count": 10, "new_count": 3}
    assert data["sources"][0]["status"] == "ok"
    assert data["sources"][0]["found"] == 12


def test_export_html_report_writes_review_dashboard(tmp_path: Path) -> None:
    output = tmp_path / "latest.html"
    tender = replace(
        make_tender("matched"),
        document_matches=["мфу", "симферополь"],
        delivery_region_evidence="notice.pdf: regions=симферополь; terms=мфу",
        source_confidence=0.9,
    )
    report = SourceFetchResult(
        health=[
            SourceHealth(
                source="ImportFolderSource",
                status="ok",
                found=1,
                elapsed_seconds=0.1,
            )
        ]
    )

    export_html_report(
        [tender],
        output,
        source_report=report,
        raw_count=1,
        unique_count=1,
        new_count=1,
    )

    html = output.read_text(encoding="utf-8")
    assert "<!doctype html>" in html
    assert "Поставка МФУ" in html
    assert "https://example.test/tender-1/" in html
    assert "notice.pdf" in html
    assert "ImportFolderSource" in html
