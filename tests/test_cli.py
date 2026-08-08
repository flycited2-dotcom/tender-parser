import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from tender_parser.cli import _all_keywords, build_default_source, build_source_for_profile, run
from tender_parser.models import TenderRecord
from tender_parser.sources.b2b_center import B2BCenterSource
from tender_parser.sources.composite import CompositeSource
from tender_parser.sources.eat import EatIntegrationSource
from tender_parser.sources.eis import EisZakupkiSource
from tender_parser.sources.etp_gpb import EtpGpbApiSource
from tender_parser.sources.rostender import RostenderSource
from tender_parser.sources.rts import RtsPublicSource, SourceFetchError
from tender_parser.sources.rts_cabinet import RtsCabinetBrowserSource
from tender_parser.sources.tender_pro import TenderProSource
from tender_parser.sources.torgi82 import Torgi82Source


class FakeSource:
    def fetch_keywords(self, keywords: list[str]) -> list[TenderRecord]:
        return [
            TenderRecord(
                title="Поставка МФУ в Республику Крым",
                url="https://example.test/tender-1/",
                source="fake",
                tender_number="1",
                customer="Заказчик",
                region="Республика Крым",
                price=45_000.0,
                deadline=datetime(2026, 5, 25, 10, 0),
                raw_text="Поставка МФУ в Республику Крым",
            )
        ]


class EmptySource:
    def fetch_keywords(self, keywords: list[str]) -> list[TenderRecord]:
        return []


class BlockedSource:
    def fetch_keywords(self, keywords: list[str]) -> list[TenderRecord]:
        raise SourceFetchError("все источники RTS недоступны")


class ReviewSource:
    def fetch_keywords(self, keywords: list[str]) -> list[TenderRecord]:
        return [
            TenderRecord(
                title="Поставка МФУ в Республику Крым",
                url="https://example.test/tender-1/",
                source="fake",
                tender_number="1",
                customer="Заказчик",
                region="Республика Крым",
                price=45_000.0,
                deadline=datetime(2026, 5, 25, 10, 0),
                raw_text="Поставка МФУ в Республику Крым",
            ),
            TenderRecord(
                title="Оказание услуг по техническому обслуживанию кондиционеров",
                url="https://example.test/tender-2/",
                source="fake",
                tender_number="2",
                customer="Заказчик",
                price=120_000.0,
                deadline=datetime(2026, 5, 25, 10, 0),
                raw_text="Оказание услуг по техническому обслуживанию кондиционеров",
            ),
        ]


class DuplicateSource:
    def fetch_keywords(self, keywords: list[str]) -> list[TenderRecord]:
        return [
            TenderRecord(
                title="Поставка МФУ в Республику Крым",
                url="https://rostender.info/tender-1/",
                source="rostender",
                tender_number="rostender-1",
                customer="Заказчик",
                region="Крым республика",
                price=45_000.0,
                deadline=datetime(2026, 5, 25, 10, 0),
                raw_text="Поставка МФУ в Республику Крым",
            ),
            TenderRecord(
                title="Поставка МФУ в Республику Крым",
                url="https://zakupki.gov.ru/tender-1/",
                source="eis-zakupki",
                tender_number="eis-1",
                region="Республика Крым",
                price=45_000.0,
                deadline=datetime(2026, 5, 25, 23, 59),
                raw_text="Поставка МФУ в Республику Крым",
            ),
        ]


def test_run_dry_mode_creates_export_dirs(tmp_path: Path) -> None:
    result = run(["--dry-run", "--base-dir", str(tmp_path)])

    assert result == 0
    assert (tmp_path / "data").is_dir()
    assert (tmp_path / "exports").is_dir()


def test_check_env_returns_one_when_eat_config_missing(tmp_path: Path, monkeypatch, capsys) -> None:
    monkeypatch.delenv("EAT_API_TOKEN", raising=False)
    monkeypatch.delenv("EAT_EXT_SYSTEM", raising=False)

    result = run(["check-env", "--base-dir", str(tmp_path)])

    output = capsys.readouterr().out
    assert result == 1
    assert "EAT_API_TOKEN: missing" in output
    assert "EAT_EXT_SYSTEM: missing" in output


def test_check_env_loads_dotenv_and_returns_zero_when_eat_config_present(
    tmp_path: Path,
    monkeypatch,
    capsys,
) -> None:
    monkeypatch.delenv("EAT_API_TOKEN", raising=False)
    monkeypatch.delenv("EAT_EXT_SYSTEM", raising=False)
    (tmp_path / ".env").write_text(
        "EAT_API_TOKEN=secret-token\nEAT_EXT_SYSTEM=EXT-CRM\n",
        encoding="utf-8",
    )

    result = run(["check-env", "--base-dir", str(tmp_path)])

    output = capsys.readouterr().out
    assert result == 0
    assert "EAT_API_TOKEN: configured" in output
    assert "secret-token" not in output


def test_all_keywords_includes_broad_aliases_and_regions() -> None:
    keywords = _all_keywords()

    assert "оргтехника" in keywords
    assert "офисная техника" in keywords
    assert "Симферополь" in keywords
    assert "Республика Крым" in keywords


def test_all_keywords_includes_expanded_network_and_electrical_terms() -> None:
    keywords = _all_keywords()

    assert "сетевое оборудование" in keywords
    assert "точка доступа" in keywords
    assert "электротехническая продукция" in keywords
    assert "разъединитель" in keywords
    assert "изолятор" in keywords
    assert "трансформатор тока" in keywords


def test_build_default_source_uses_composite_source() -> None:
    source = build_default_source()

    assert isinstance(source, CompositeSource)
    first_layer = source.sources[0]
    assert isinstance(first_layer, CompositeSource)
    assert isinstance(first_layer.sources[0], EtpGpbApiSource)
    assert isinstance(first_layer.sources[1], TenderProSource)
    assert isinstance(first_layer.sources[2], Torgi82Source)
    assert isinstance(first_layer.sources[3], B2BCenterSource)
    assert isinstance(first_layer.sources[4], EatIntegrationSource)
    assert isinstance(first_layer.sources[5], EisZakupkiSource)
    assert isinstance(first_layer.sources[6], RostenderSource)
    assert isinstance(first_layer.sources[7], RtsPublicSource)
    assert len(source.sources) == 1


def test_build_fast_profile_includes_eis_after_native_tls_fix() -> None:
    source = build_source_for_profile("fast")

    assert isinstance(source, CompositeSource)
    names = [item.__class__.__name__ for item in source.sources[0].sources]
    assert names == [
        "EtpGpbApiSource",
        "TenderProSource",
        "Torgi82Source",
        "B2BCenterSource",
        "EatIntegrationSource",
        "EisZakupkiSource",
        "RostenderSource",
    ]


def test_build_rts_profile_runs_only_rts_source() -> None:
    source = build_source_for_profile("rts")

    assert isinstance(source, CompositeSource)
    assert len(source.sources) == 1
    assert isinstance(source.sources[0], RtsPublicSource)


def test_build_rts_cabinet_profile_runs_only_cabinet_source() -> None:
    source = build_source_for_profile("rts-cabinet")

    assert isinstance(source, CompositeSource)
    assert len(source.sources) == 1
    assert isinstance(source.sources[0], RtsCabinetBrowserSource)


def test_run_local_profile_uses_imports_without_live_sources(tmp_path: Path) -> None:
    imports_dir = tmp_path / "imports"
    imports_dir.mkdir()
    (imports_dir / "manual.csv").write_text(
        "Название;Ссылка;Номер;Заказчик;Регион;Сумма;Срок подачи;Источник\n"
        "Поставка принтеров;https://example.test/local;LOCAL-1;Администрация;Севастополь;55000;25.05.2026 10:00;Manual\n",
        encoding="utf-8",
    )

    result = run(
        ["--profile", "local", "--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"],
    )

    latest = json.loads((tmp_path / "exports" / "latest.json").read_text(encoding="utf-8"))
    run_report = json.loads((tmp_path / "exports" / "run_report.json").read_text(encoding="utf-8"))

    assert result == 0
    assert latest["count"] == 1
    assert latest["items"][0]["source"] == "import-manual"
    assert [source["source"] for source in run_report["sources"]] == ["ImportFolderSource"]


def test_run_rts_profile_preserves_diagnostic_report_when_all_endpoints_fail(tmp_path: Path) -> None:
    result = run(
        ["--profile", "rts", "--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"],
        source=CompositeSource([BlockedSource()]),
    )

    run_report = json.loads((tmp_path / "exports" / "run_report.json").read_text(encoding="utf-8"))

    assert result == 0
    assert run_report["summary"]["raw_count"] == 0
    assert run_report["sources"][0]["source"] == "BlockedSource"
    assert run_report["sources"][0]["status"] == "error"


def test_run_with_fake_source_creates_database_and_exports(tmp_path: Path) -> None:
    result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"],
        source=FakeSource(),
    )

    assert result == 0
    assert (tmp_path / "data" / "tenders.db").exists()
    assert (tmp_path / "exports" / "latest.json").exists()
    assert (tmp_path / "exports" / "latest.html").exists()
    assert (tmp_path / "exports" / "new_tenders.json").exists()
    assert (tmp_path / "exports" / "notification.txt").exists()
    assert (tmp_path / "exports" / "run_report.json").exists()
    excel_path = next((tmp_path / "exports").glob("tenders_*.xlsx"))
    workbook = load_workbook(excel_path)
    assert workbook.sheetnames == ["Дашборд", "Новые", "Горячие", "На проверку", "Широкий хвост", "Отсеянные"]
    assert "Поставка МФУ" in (tmp_path / "exports" / "latest.html").read_text(encoding="utf-8")
    assert "Поставка МФУ" in (tmp_path / "exports" / "notification.txt").read_text(
        encoding="utf-8"
    )


def test_run_preserves_new_queue_when_export_fails(tmp_path: Path) -> None:
    import pytest

    (tmp_path / "exports").mkdir(parents=True)
    (tmp_path / "exports" / "latest.json").mkdir()  # каталог вместо файла блокирует запись

    with pytest.raises(OSError):
        run(["--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"], source=FakeSource())

    (tmp_path / "exports" / "latest.json").rmdir()
    result = run(["--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"], source=FakeSource())

    assert result == 0
    data = json.loads((tmp_path / "exports" / "new_tenders.json").read_text(encoding="utf-8"))
    assert data["count"] == 1  # тендер не потерян из CRM-очереди после упавшего экспорта


def test_run_rejects_invalid_now(tmp_path: Path, capsys) -> None:
    result = run(["--base-dir", str(tmp_path), "--now", "вчера"], source=FakeSource())

    assert result == 2


def test_run_exports_only_current_run_matches(tmp_path: Path) -> None:
    first_result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"],
        source=FakeSource(),
    )
    second_result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-20T12:00:00"],
        source=EmptySource(),
    )

    assert first_result == 0
    assert second_result == 0
    assert '"count": 0' in (tmp_path / "exports" / "latest.json").read_text(encoding="utf-8")


def test_run_exports_only_first_seen_actionable_tenders(tmp_path: Path) -> None:
    first_result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"],
        source=FakeSource(),
    )
    first_new = (tmp_path / "exports" / "new_tenders.json").read_text(encoding="utf-8")

    second_result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-20T12:00:00"],
        source=FakeSource(),
    )
    second_new = (tmp_path / "exports" / "new_tenders.json").read_text(encoding="utf-8")

    assert first_result == 0
    assert second_result == 0
    assert '"count": 1' in first_new
    assert '"count": 0' in second_new


def test_run_restores_known_price_and_deadline_before_filtering(tmp_path: Path) -> None:
    first_result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"],
        source=FakeSource(),
    )

    class PartialRepeatSource:
        def fetch_keywords(self, keywords: list[str]) -> list[TenderRecord]:
            original = FakeSource().fetch_keywords(keywords)[0]
            return [
                TenderRecord(
                    title=original.title,
                    url=original.url,
                    source=original.source,
                    tender_number=original.tender_number,
                    customer=original.customer,
                    region=original.region,
                    price=None,
                    deadline=None,
                    raw_text=original.raw_text,
                )
            ]

    second_result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-20T12:00:00"],
        source=PartialRepeatSource(),
    )
    latest = json.loads((tmp_path / "exports" / "latest.json").read_text(encoding="utf-8"))

    assert first_result == 0
    assert second_result == 0
    assert latest["count"] == 1
    assert latest["items"][0]["price"] == 45_000.0
    assert latest["items"][0]["deadline"] == "2026-05-25T10:00:00"


def test_run_keeps_existing_exports_when_source_is_blocked(tmp_path: Path) -> None:
    exports_dir = tmp_path / "exports"
    exports_dir.mkdir()
    latest = exports_dir / "latest.json"
    latest.write_text("old report", encoding="utf-8")

    result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"],
        source=BlockedSource(),
    )

    assert result == 2
    assert latest.read_text(encoding="utf-8") == "old report"


def test_run_exports_review_items_for_manual_check(tmp_path: Path) -> None:
    result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"],
        source=ReviewSource(),
    )

    latest = (tmp_path / "exports" / "latest.json").read_text(encoding="utf-8")

    assert result == 0
    assert '"count": 2' in latest
    assert '"filter_status": "matched"' in latest
    assert '"filter_status": "review"' in latest


def test_run_merges_cross_source_duplicates(tmp_path: Path) -> None:
    result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"],
        source=DuplicateSource(),
    )

    latest = (tmp_path / "exports" / "latest.json").read_text(encoding="utf-8")

    assert result == 0
    assert '"count": 1' in latest
    assert '"source": "eis-zakupki"' in latest


def test_run_adds_import_folder_records_to_current_report(tmp_path: Path) -> None:
    imports_dir = tmp_path / "imports"
    imports_dir.mkdir()
    (imports_dir / "manual.csv").write_text(
        "Название;Ссылка;Номер;Заказчик;Регион;Сумма;Срок подачи;Источник\n"
        "Поставка МФУ;https://example.test/imported;IMP-1;Администрация;Симферополь;45000;25.05.2026 10:00;Manual\n",
        encoding="utf-8",
    )

    result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"],
        source=EmptySource(),
    )

    latest = json.loads((tmp_path / "exports" / "latest.json").read_text(encoding="utf-8"))
    run_report = json.loads((tmp_path / "exports" / "run_report.json").read_text(encoding="utf-8"))

    assert result == 0
    assert latest["count"] == 1
    assert latest["items"][0]["source"] == "import-manual"
    assert latest["items"][0]["detail_status"] == "imported"
    assert any(source["source"] == "ImportFolderSource" and source["found"] == 1 for source in run_report["sources"])


def test_run_enriches_records_from_documents_before_filtering(tmp_path: Path) -> None:
    documents_dir = tmp_path / "documents"
    documents_dir.mkdir()
    (documents_dir / "specification.txt").write_text(
        "Адрес поставки: г. Севастополь. Требуется поставка многофункциональных устройств.",
        encoding="utf-8",
    )

    result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"],
        source=EmptySource(),
    )

    assert result == 0
    assert json.loads((tmp_path / "exports" / "latest.json").read_text(encoding="utf-8"))["count"] == 0

    class DocumentOnlySource:
        def fetch_keywords(self, keywords: list[str]) -> list[TenderRecord]:
            return [
                TenderRecord(
                    title="Поставка оргтехники",
                    url="https://example.test/document-only",
                    source="fake",
                    tender_number="DOC-1",
                    price=55_000,
                    deadline=datetime(2026, 5, 25, 10, 0),
                )
            ]

    second_result = run(
        ["--base-dir", str(tmp_path), "--now", "2026-05-19T12:00:00"],
        source=DocumentOnlySource(),
    )

    latest = json.loads((tmp_path / "exports" / "latest.json").read_text(encoding="utf-8"))

    assert second_result == 0
    assert latest["count"] == 1
    assert latest["items"][0]["detail_status"] == "enriched"
    assert "севастополь" in latest["items"][0]["document_matches"]
    assert "specification.txt" in latest["items"][0]["delivery_region_evidence"]
