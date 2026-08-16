import json
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from tender_parser.cli import run
from tender_parser.tender_case import (
    LineItem,
    ProductOffer,
    TenderCase,
    calculate_case,
)


def _offer(*, price: str = "6000", status: str = "exact", evidence: str = "Паспорт, стр. 2") -> ProductOffer:
    return ProductOffer(
        line_id="1",
        supplier="Основной поставщик",
        sku="SKU-1",
        product_name="Тестовый товар",
        unit_cost_gross=Decimal(price),
        compliance_status=status,  # type: ignore[arg-type]
        stock="В наличии",
        lead_days=7,
        evidence=evidence,
    )


def _case(*, nmck: str = "15000", planned_bid: str | None = None, region: str = "Симферополь") -> TenderCase:
    return TenderCase(
        case_id="case-1",
        title="Поставка товара",
        region=region,
        nmck=Decimal(nmck),
        planned_bid=Decimal(planned_bid) if planned_bid else None,
    )


def _items() -> list[LineItem]:
    return [LineItem(line_id="1", name="Товар", quantity=Decimal("1"), required_specs="Мощность не менее 1 кВт")]


def test_calculate_case_uses_30_15_and_12_percent_thresholds() -> None:
    economics = calculate_case(_case(), _items(), [_offer()], [])

    assert economics.procurement_gross == Decimal("6000.00")
    assert economics.target_price == Decimal("7800.00")
    assert economics.viable_price == Decimal("6900.00")
    assert economics.hard_floor_price == Decimal("6720.00")
    assert economics.target_discount_from_nmck == Decimal("0.4800")
    assert economics.viable_discount_from_nmck == Decimal("0.5400")
    assert economics.hard_floor_discount_from_nmck == Decimal("0.5520")
    assert economics.decision == "ready"

    review = calculate_case(_case(planned_bid="6800"), _items(), [_offer()], [])
    assert review.decision == "manual_review"
    assert "12% до 15%" in review.decision_reason

    stopped = calculate_case(_case(planned_bid="6600"), _items(), [_offer()], [])
    assert stopped.decision == "stop"


def test_missing_or_conditional_product_cannot_be_approved_automatically() -> None:
    missing = calculate_case(_case(), _items(), [], [])
    assert missing.decision == "blocked"

    conditional = calculate_case(
        _case(),
        _items(),
        [_offer(status="conditional")],
        [],
    )
    # Условный товар не выбирается автоматически.
    assert conditional.decision == "blocked"

    manually_selected = ProductOffer(**{**_offer(status="conditional").__dict__, "selected": True})
    manual = calculate_case(_case(), _items(), [manually_selected], [])
    assert manual.decision == "manual_review"


def test_new_regions_require_confirmed_delivery_cost() -> None:
    economics = calculate_case(_case(region="Запорожская область"), _items(), [_offer()], [])

    assert economics.decision == "manual_review"
    assert any("стоимость доставки" in risk for risk in economics.risks)


def test_entity_scenarios_keep_input_vat_in_ip_cost() -> None:
    economics = calculate_case(_case(nmck="13000"), _items(), [_offer(price="10000")], [])
    ooo, ip_10, ip_15 = economics.entity_scenarios

    assert ooo.procurement_cost == Decimal("8196.72")
    assert ooo.profit_before_income_tax == Decimal("2459.02")
    assert ip_10.procurement_cost == Decimal("10000.00")
    assert ip_10.profit_before_income_tax == Decimal("3000.00")
    assert ip_10.estimated_tax == Decimal("300.00")
    assert ip_10.minimum_tax_reference == Decimal("130.00")
    assert ip_10.profit_after_estimated_tax == Decimal("2700.00")
    assert ip_15.estimated_tax == Decimal("450.00")
    assert "нарастающим итогом" in ip_10.note.lower()


def test_case_cli_creates_templates_and_excel_report(tmp_path: Path) -> None:
    assert run(["case-init", "--base-dir", str(tmp_path), "--case-id", "test-77", "--title", "Поставка МФУ"]) == 0
    case_dir = tmp_path / "cases" / "test-77"
    assert (case_dir / "case.json").exists()

    payload = json.loads((case_dir / "case.json").read_text(encoding="utf-8"))
    payload.update({"nmck": 15000, "region": "Симферополь"})
    (case_dir / "case.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    (case_dir / "items.csv").write_text(
        "line_id;name;quantity;unit;required_specs;mandatory\n1;МФУ;1;шт.;A4;yes\n",
        encoding="utf-8-sig",
    )
    (case_dir / "offers.csv").write_text(
        "line_id;supplier;sku;product_name;unit_cost_gross;compliance_status;selected;stock;lead_days;vat_rate;source_url;evidence;notes\n"
        "1;Поставщик;SKU-1;МФУ;6000;exact;yes;В наличии;7;0.22;https://example.test;Паспорт стр. 2;\n",
        encoding="utf-8-sig",
    )

    assert run(["case-report", "--base-dir", str(tmp_path), "--case-id", "test-77"]) == 0
    report = case_dir / "output" / "case_report.xlsx"
    assert report.exists()
    workbook = load_workbook(report, read_only=True)
    try:
        assert workbook.sheetnames == [
            "Решение",
            "Позиции",
            "Предложения",
            "Расходы",
            "ООО и ИП",
            "Торги",
            "Риски",
            "Вопросы",
            "Документы",
            "Комплект заявки",
        ]
        assert workbook["Решение"]["B3"].value == "МОЖНО РАССМАТРИВАТЬ УЧАСТИЕ"
    finally:
        workbook.close()
