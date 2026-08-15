import json
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from tender_parser.exporters.excel import DATE_FORMAT, export_excel
from tender_parser.exporters.html_report import export_html_report
from tender_parser.exporters.json_exporter import export_json
from tender_parser.models import TenderRecord
from tender_parser.run_report import SourceFetchResult
from tender_parser.storage import TenderStorage


def enriched_tender() -> TenderRecord:
    return TenderRecord(
        title="Поставка МФУ в Симферополь",
        url="https://example.test/tender-1/",
        source="import",
        tender_number="1",
        region="Симферополь",
        price=45_000.0,
        deadline=datetime(2026, 7, 1, 10, 0),
        filter_status="matched",
        review_priority="hot",
        detail_status="enriched",
        document_matches=["мфу", "Симферополь"],
        delivery_region_evidence="Адрес поставки: г. Симферополь",
        source_confidence=0.85,
        official_number="0174100000626000005",
        official_url=(
            "https://zakupki.gov.ru/epz/order/notice/ea20/view/common-info.html"
            "?regNumber=0174100000626000005"
        ),
        official_source="ЕИС",
        platform_number="SBER-42",
        platform_url="https://utp.sberbank-ast.ru/Trade/NBT/PurchaseView/42/0/0/0",
        procurement_law="44-ФЗ",
        resolution_method="rostender-meta",
        resolution_confidence=0.98,
    )


def test_json_export_includes_enrichment_fields(tmp_path: Path) -> None:
    output = tmp_path / "latest.json"

    export_json([enriched_tender()], output)

    item = json.loads(output.read_text(encoding="utf-8"))["items"][0]
    assert item["detail_status"] == "enriched"
    assert item["document_matches"] == ["мфу", "Симферополь"]
    assert item["delivery_region_evidence"] == "Адрес поставки: г. Симферополь"
    assert item["source_confidence"] == 0.85
    assert item["official_number"] == "0174100000626000005"
    assert item["official_url"].startswith("https://zakupki.gov.ru/")
    assert item["official_source"] == "ЕИС"
    assert item["platform_number"] == "SBER-42"
    assert item["platform_url"].startswith("https://utp.sberbank-ast.ru/")
    assert item["procurement_law"] == "44-ФЗ"
    assert item["resolution_method"] == "rostender-meta"
    assert item["resolution_confidence"] == 0.98
    assert item["tender_number"] == "1"
    assert item["url"] == "https://example.test/tender-1/"


def test_excel_export_includes_enrichment_columns(tmp_path: Path) -> None:
    output = tmp_path / "tenders.xlsx"

    export_excel([enriched_tender()], [], [], [], output)

    workbook = load_workbook(output)
    headers = [cell.value for cell in workbook["Горячие"][1]]
    row = [cell.value for cell in workbook["Горячие"][2]]
    assert "detail_status" in headers
    assert "document_matches" in headers
    assert "delivery_region_evidence" in headers
    assert "source_confidence" in headers
    assert row[headers.index("detail_status")] == "enriched"
    assert row[headers.index("document_matches")] == "мфу; Симферополь"
    assert row[headers.index("source_confidence")] == 0.85
    assert row[headers.index("официальный номер")] == "0174100000626000005"
    assert row[headers.index("официальный источник")] == "ЕИС"
    assert row[headers.index("номер на площадке")] == "SBER-42"
    assert row[headers.index("закон")] == "44-ФЗ"
    assert row[headers.index("способ определения")] == "rostender-meta"
    assert row[headers.index("уверенность определения")] == 0.98
    official_number_cell = workbook["Горячие"].cell(
        2, headers.index("официальный номер") + 1
    )
    assert official_number_cell.value == "0174100000626000005"
    assert official_number_cell.data_type == "s"
    assert official_number_cell.number_format == "@"
    deadline_cell = workbook["Горячие"].cell(2, headers.index("срок_подачи") + 1)
    assert deadline_cell.value == datetime(2026, 7, 1, 10, 0)
    assert deadline_cell.data_type == "d"
    assert deadline_cell.number_format == DATE_FORMAT
    assert (
        workbook["Горячие"].cell(2, headers.index("прямая ссылка") + 1).hyperlink.target
        == enriched_tender().official_url
    )
    assert (
        workbook["Горячие"]
        .cell(2, headers.index("ссылка на площадку") + 1)
        .hyperlink.target
        == enriched_tender().platform_url
    )
    # Aggregator/source provenance remains available and clickable on the title.
    assert row[headers.index("номер")] == "1"
    assert workbook["Горячие"].cell(2, headers.index("название") + 1).hyperlink.target == (
        enriched_tender().url
    )


def test_html_export_links_to_official_source_and_keeps_aggregator_url(tmp_path: Path) -> None:
    output = tmp_path / "latest.html"
    tender = enriched_tender()

    export_html_report(
        [tender],
        output,
        source_report=SourceFetchResult(),
        raw_count=1,
        unique_count=1,
        new_count=1,
    )

    html = output.read_text(encoding="utf-8")
    assert tender.url in html
    assert tender.official_url in html
    assert tender.platform_url in html
    assert tender.official_number in html
    assert "44-ФЗ" in html


def test_storage_round_trips_enrichment_fields(tmp_path: Path) -> None:
    storage = TenderStorage(tmp_path / "tenders.db")
    tender = enriched_tender()

    storage.upsert_many([tender])

    loaded = storage.fetch_by_status("matched")[0]
    assert loaded.detail_status == "enriched"
    assert loaded.document_matches == ["мфу", "Симферополь"]
    assert loaded.delivery_region_evidence == "Адрес поставки: г. Симферополь"
    assert loaded.source_confidence == 0.85
    assert loaded.official_number == "0174100000626000005"
    assert loaded.official_url == tender.official_url
    assert loaded.official_source == "ЕИС"
    assert loaded.platform_number == "SBER-42"
    assert loaded.platform_url == tender.platform_url
    assert loaded.procurement_law == "44-ФЗ"
    assert loaded.resolution_method == "rostender-meta"
    assert loaded.resolution_confidence == 0.98


def test_storage_migrates_legacy_database_for_enrichment_fields(tmp_path: Path) -> None:
    db_path = tmp_path / "tenders.db"
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE tenders (
                unique_key TEXT PRIMARY KEY,
                title TEXT NOT NULL,
                url TEXT NOT NULL,
                source TEXT NOT NULL,
                tender_number TEXT,
                customer TEXT,
                region TEXT,
                price REAL,
                deadline TEXT,
                status TEXT,
                published_at TEXT,
                discovered_at TEXT,
                last_seen_at TEXT,
                raw_text TEXT,
                category TEXT,
                include_reason TEXT,
                exclude_reason TEXT,
                filter_status TEXT NOT NULL,
                match_confidence TEXT,
                review_priority TEXT
            )
            """
        )

    TenderStorage(db_path)

    with sqlite3.connect(db_path) as conn:
        columns = {row[1] for row in conn.execute("PRAGMA table_info(tenders)")}
    assert "detail_status" in columns
    assert "document_matches" in columns
    assert "delivery_region_evidence" in columns
    assert "source_confidence" in columns
    assert {
        "official_number",
        "official_url",
        "official_source",
        "platform_number",
        "platform_url",
        "procurement_law",
        "resolution_method",
        "resolution_confidence",
    } <= columns
