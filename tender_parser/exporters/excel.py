from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Mapping

from math import ceil, isfinite
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from tender_parser.models import TenderRecord
from tender_parser.run_report import SourceHealth


HEADERS = [
    "приоритет",
    "новый",
    "дней_до_срока",
    "название",
    "регион",
    "сумма",
    "срок_подачи",
    "заказчик",
    "категория",
    "уверенность",
    "мой_выбор",
    "комментарий",
    "номер",
    "источник",
    "статус",
    "дата_обнаружения",
    "причина_включения",
    "причина_исключения",
    "detail_status",
    "document_matches",
    "delivery_region_evidence",
    "source_confidence",
]

COLUMN_WIDTHS = {
    "приоритет": 12,
    "новый": 7,
    "дней_до_срока": 9,
    "название": 55,
    "регион": 20,
    "сумма": 16,
    "срок_подачи": 17,
    "заказчик": 32,
    "категория": 26,
    "уверенность": 15,
    "мой_выбор": 12,
    "комментарий": 24,
    "номер": 16,
    "источник": 18,
    "статус": 16,
    "дата_обнаружения": 17,
    "причина_включения": 45,
    "причина_исключения": 35,
    "detail_status": 16,
    "document_matches": 24,
    "delivery_region_evidence": 30,
    "source_confidence": 12,
}

PRIORITY_ORDER = {"hot": 0, "review": 1, "wide": 2, "excluded": 3, None: 4}
PRIORITY_LABELS = {"hot": "Горячий", "review": "На проверку", "wide": "Широкий хвост", "excluded": "Отсеян"}

# Светлые тинты статусной палитры: смысл всегда продублирован текстом приоритета.
PRIORITY_FILLS = {
    "hot": "FFFBE0D6",       # serious #ec835a, тинт 25%
    "review": "FFFEECCA",    # warning #fab219, тинт 25%
    "wide": "FFDCECFF",      # информационный голубой из HTML-отчета
    "excluded": "FFECEFF1",  # нейтральный серый
}
DEADLINE_URGENT_FILL = "FFF5CECE"   # critical #d03b3b, тинт 25%: 2 дня и меньше
DEADLINE_SOON_FILL = "FFFEECCA"     # warning: неделя и меньше
HEADER_FILL = "FF1A1A19"
CHART_BAR_COLOR = "2A78D6"          # sequential blue, одна серия
KPI_HOT_COLOR = "FFD03B3B"
HEALTH_FILLS = {
    "ok": "FFCEEDCE",
    "partial": "FFFEECCA",
    "empty": "FFECEFF1",
    "skipped": "FFECEFF1",
}
HEALTH_BAD_FILL = "FFF5CECE"

TAB_COLORS = {
    "Дашборд": "2A78D6",
    "Мой отбор": "104281",
    "Новые": "0CA30C",
    "Горячие": "EC835A",
    "На проверку": "FAB219",
    "Широкий хвост": "86B6EF",
    "Отсеянные": "9AA0A6",
}

DATA_SHEETS = {"Мой отбор", "Новые", "Горячие", "На проверку", "Широкий хвост", "Отсеянные"}
# Порядок чтения отметок: листы данных раньше агрегата «Мой отбор»,
# чтобы правка пользователя на «Горячих» побеждала устаревшую копию.
DATA_SHEET_READ_ORDER = ["Новые", "Горячие", "На проверку", "Широкий хвост", "Отсеянные", "Мой отбор"]
PICKED_CHOICES = ("✅", "🤔")


def _safe(value: object) -> object:
    """Excel запрещает управляющие символы — они приходят из вёрстки площадок."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _safe_price(price: float | None) -> float | None:
    if price is None or not isfinite(price):
        return None
    return price


def _normalize_number(value: object) -> str:
    # Excel хранит перевбитые номера как float: 123.0 должен совпасть с "123".
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "").strip()

SelectionKey = tuple[str, str]
SelectionValue = tuple[str, str]

CHOICE_VALIDATION = '"✅ Беру,🤔 Думаю,❌ Мимо"'
PRICE_FORMAT = '#,##0.00 "₽"'
DATE_FORMAT = "DD.MM.YYYY HH:MM"

HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=10)
LINK_FONT = Font(color="FF0B63CE", underline="single")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")

_GRID_SIDE = Side(style="thin", color="FFD8DDE3")
GRID_BORDER = Border(left=_GRID_SIDE, right=_GRID_SIDE, top=_GRID_SIDE, bottom=_GRID_SIDE)

# Колонки с переносом текста: по ним считается высота строки (ширина уменьшена на поля).
WRAP_COLUMNS = {
    "название": 53,
    "причина_включения": 43,
    "причина_исключения": 33,
    "delivery_region_evidence": 28,
}
ROW_LINE_HEIGHT = 13.5
MAX_ROW_LINES = 5


def _row_lines(tender: TenderRecord) -> int:
    values = {
        "название": tender.title,
        "причина_включения": tender.include_reason,
        "причина_исключения": tender.exclude_reason,
        "delivery_region_evidence": tender.delivery_region_evidence,
    }
    lines = 1
    for header, width in WRAP_COLUMNS.items():
        text = values.get(header) or ""
        if text:
            lines = max(lines, ceil(len(text) / width))
    return min(lines, MAX_ROW_LINES)


def _format_dt(value: datetime | None) -> str:
    return value.strftime("%Y-%m-%d %H:%M") if value else ""


def _days_left(deadline: datetime | None, now: datetime) -> int | None:
    if deadline is None:
        return None
    return (deadline.date() - now.date()).days


def sort_for_review(tenders: list[TenderRecord]) -> list[TenderRecord]:
    return sorted(
        tenders,
        key=lambda tender: (
            PRIORITY_ORDER.get(tender.review_priority, 4),
            tender.deadline is None,
            tender.deadline or datetime.max,
            tender.price is None,
            -(tender.price or 0),
            -(tender.discovered_at.timestamp() if tender.discovered_at else 0),
            tender.title,
        ),
    )


def load_manual_selections(exports_dir: Path) -> dict[SelectionKey, SelectionValue]:
    """Читает отметки мой_выбор/комментарий из последнего выгруженного Excel.

    Отметки пользователя переживают перегенерацию отчета: ключ — (источник, номер).
    """
    if not exports_dir.exists():
        return {}
    # По mtime (не по имени): fallback-файлы вида _HHMMSS не должны затенять
    # более свежий основной файл. Читаем до трёх последних, свежайший побеждает.
    candidates = sorted(
        exports_dir.glob("tenders_*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True
    )[:3]
    from openpyxl import load_workbook

    selections: dict[SelectionKey, SelectionValue] = {}
    for candidate in candidates:
        try:
            workbook = load_workbook(candidate, read_only=True, data_only=True)
        except Exception:
            print(f"Не удалось прочитать отметки из {candidate.name} — файл пропущен")
            continue
        try:
            for name in DATA_SHEET_READ_ORDER:
                if name not in workbook.sheetnames:
                    continue
                sheet = workbook[name]
                rows = sheet.iter_rows(values_only=True)
                headers = [str(value or "").strip() for value in next(rows, [])]
                try:
                    source_col = headers.index("источник")
                    number_col = headers.index("номер")
                    choice_col = headers.index("мой_выбор")
                    comment_col = headers.index("комментарий")
                except ValueError:
                    continue
                for row in rows:
                    if len(row) <= max(source_col, number_col, choice_col, comment_col):
                        continue
                    choice = str(row[choice_col] or "").strip()
                    comment = str(row[comment_col] or "").strip()
                    if not choice and not comment:
                        continue
                    key = (str(row[source_col] or ""), _normalize_number(row[number_col]))
                    if key not in selections:
                        selections[key] = (choice, comment)
        finally:
            workbook.close()
    return selections


def _selection_for(
    tender: TenderRecord, selections: Mapping[SelectionKey, SelectionValue] | None
) -> SelectionValue:
    if not selections:
        return "", ""
    return selections.get((tender.source, tender.tender_number or ""), ("", ""))


def _style_header(sheet: Worksheet, columns: int, row: int = 1) -> None:
    fill = PatternFill("solid", start_color=HEADER_FILL)
    for index in range(1, columns + 1):
        cell = sheet.cell(row=row, column=index)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = GRID_BORDER


def _border_row(sheet: Worksheet, row: int, columns: int) -> None:
    for index in range(1, columns + 1):
        sheet.cell(row=row, column=index).border = GRID_BORDER


def _append_rows(
    sheet: Worksheet,
    tenders: list[TenderRecord],
    now: datetime,
    selections: Mapping[SelectionKey, SelectionValue] | None = None,
    new_keys: set[str] | None = None,
) -> None:
    sheet.append(HEADERS)
    _style_header(sheet, len(HEADERS))
    for header, width in COLUMN_WIDTHS.items():
        letter = get_column_letter(HEADERS.index(header) + 1)
        sheet.column_dimensions[letter].width = width

    choice_validation = DataValidation(type="list", formula1=CHOICE_VALIDATION, allow_blank=True)
    sheet.add_data_validation(choice_validation)
    choice_column = get_column_letter(HEADERS.index("мой_выбор") + 1)

    for tender in tenders:
        days = _days_left(tender.deadline, now)
        choice, comment = _selection_for(tender, selections)
        sheet.append(
            [
                _safe(value)
                for value in [
                    PRIORITY_LABELS.get(tender.review_priority or "", tender.review_priority or ""),
                    "🆕" if new_keys and tender.unique_key in new_keys else "",
                    days if days is not None else "",
                    tender.title,
                    tender.region or "",
                    _safe_price(tender.price),
                    tender.deadline,
                    tender.customer or "",
                    tender.category or "",
                    tender.match_confidence or "",
                    choice,
                    comment,
                    tender.tender_number or "",
                    tender.source,
                    tender.status or "",
                    _format_dt(tender.discovered_at),
                    tender.include_reason,
                    tender.exclude_reason,
                    tender.detail_status,
                    "; ".join(tender.document_matches),
                    tender.delivery_region_evidence,
                    tender.source_confidence,
                ]
            ]
        )
        row = sheet.max_row
        row_fill_color = PRIORITY_FILLS.get(tender.review_priority or "")
        row_fill = PatternFill("solid", start_color=row_fill_color) if row_fill_color else None
        for index in range(1, len(HEADERS) + 1):
            cell = sheet.cell(row=row, column=index)
            cell.alignment = TOP
            cell.border = GRID_BORDER
            if row_fill is not None:
                cell.fill = row_fill
        sheet.row_dimensions[row].height = ROW_LINE_HEIGHT * _row_lines(tender) + 3

        priority_cell = sheet.cell(row=row, column=1)
        priority_cell.font = Font(bold=True, size=10)

        days_cell = sheet.cell(row=row, column=HEADERS.index("дней_до_срока") + 1)
        if days is not None and days <= 2:
            days_cell.fill = PatternFill("solid", start_color=DEADLINE_URGENT_FILL)
            days_cell.font = Font(bold=True)
        elif days is not None and days <= 7:
            days_cell.fill = PatternFill("solid", start_color=DEADLINE_SOON_FILL)

        title_cell = sheet.cell(row=row, column=HEADERS.index("название") + 1)
        title_cell.alignment = WRAP_TOP
        if tender.url:
            title_cell.hyperlink = tender.url
            title_cell.font = LINK_FONT

        price_cell = sheet.cell(row=row, column=HEADERS.index("сумма") + 1)
        price_cell.number_format = PRICE_FORMAT
        deadline_cell = sheet.cell(row=row, column=HEADERS.index("срок_подачи") + 1)
        deadline_cell.number_format = DATE_FORMAT
        # Номер как текст: Excel не должен превращать 19-значные номера в float.
        sheet.cell(row=row, column=HEADERS.index("номер") + 1).number_format = "@"
        for header in ("причина_включения", "причина_исключения", "delivery_region_evidence"):
            sheet.cell(row=row, column=HEADERS.index(header) + 1).alignment = WRAP_TOP

    last_row = max(sheet.max_row, 2)
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"
    sheet.freeze_panes = "E2"
    choice_validation.add(f"{choice_column}2:{choice_column}{last_row}")


def _count_by(tenders: list[TenderRecord], key) -> list[tuple[str, int]]:
    counts: dict[str, int] = {}
    for tender in tenders:
        name = key(tender) or "не указано"
        counts[name] = counts.get(name, 0) + 1
    return sorted(counts.items(), key=lambda item: item[1], reverse=True)


def _add_bar_chart(sheet: Worksheet, title: str, first_row: int, last_row: int, anchor_row: int) -> int:
    """Горизонтальный bar со значениями на барах и именами на оси; возвращает нижнюю строку."""
    chart = BarChart()
    chart.type = "bar"
    chart.title = title
    chart.legend = None
    chart.height = max(5.0, 1.0 * (last_row - first_row + 1) + 2.5)
    chart.width = 16
    data = Reference(sheet, min_col=2, min_row=first_row - 1, max_row=last_row)
    categories = Reference(sheet, min_col=1, min_row=first_row, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    labels = DataLabelList()
    labels.showVal = True
    labels.showSerName = False
    labels.showCatName = False
    labels.showLegendKey = False
    labels.showPercent = False
    chart.dataLabels = labels
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    series = chart.series[0]
    series.graphicalProperties.solidFill = CHART_BAR_COLOR
    sheet.add_chart(chart, f"D{anchor_row}")
    # ~0.53 см на строку Excel по умолчанию
    return anchor_row + ceil(chart.height / 0.53) + 1


def _dashboard_section(sheet: Worksheet, row: int, title: str) -> int:
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=12)
    return row + 1


def _build_dashboard(
    sheet: Worksheet,
    *,
    hot: list[TenderRecord],
    review: list[TenderRecord],
    wide: list[TenderRecord],
    excluded: list[TenderRecord],
    new_count: int | None,
    now: datetime,
    source_health: list[SourceHealth] | None,
) -> None:
    sheet.column_dimensions["A"].width = 55
    for letter in "BCDEF":
        sheet.column_dimensions[letter].width = 14
    sheet.sheet_view.showGridLines = False

    title = sheet.cell(row=1, column=1, value="Тендеры — панель управления")
    title.font = Font(bold=True, size=18)
    generated = sheet.cell(row=2, column=1, value=f"Сформировано: {now.strftime('%d.%m.%Y %H:%M')}")
    generated.font = Font(color="FF6B7280")

    actionable = hot + review + wide
    kpis = [
        ("Горячие", len(hot)),
        ("На проверку", len(review)),
        ("Широкий хвост", len(wide)),
        ("Отсеянные", len(excluded)),
        ("Всего карточек", len(actionable) + len(excluded)),
    ]
    if new_count is not None:
        kpis.insert(3, ("Новые для CRM", new_count))
    for index, (label, value) in enumerate(kpis, start=1):
        label_cell = sheet.cell(row=4, column=index, value=label)
        label_cell.font = Font(color="FF6B7280", size=10)
        value_cell = sheet.cell(row=5, column=index, value=value)
        value_cell.font = Font(bold=True, size=16, color=KPI_HOT_COLOR if label == "Горячие" else None)

    row = 7
    row = _dashboard_section(sheet, row, "Actionable по источникам")
    header_row = row
    sheet.cell(row=row, column=1, value="Источник")
    sheet.cell(row=row, column=2, value="Карточек")
    _style_header(sheet, 2, row=row)
    sources = _count_by(actionable, lambda tender: tender.source)
    for name, count in sources:
        row += 1
        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=2, value=count)
        _border_row(sheet, row, 2)
    if sources:
        chart_bottom = _add_bar_chart(sheet, "Actionable по источникам", header_row + 1, row, header_row)
        row = max(row, chart_bottom)

    row += 2
    row = _dashboard_section(sheet, row, "Actionable по категориям")
    header_row = row
    sheet.cell(row=row, column=1, value="Категория")
    sheet.cell(row=row, column=2, value="Карточек")
    _style_header(sheet, 2, row=row)
    categories = _count_by(actionable, lambda tender: tender.category)
    for name, count in categories:
        row += 1
        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=2, value=count)
        _border_row(sheet, row, 2)
    if categories:
        chart_bottom = _add_bar_chart(sheet, "Actionable по категориям", header_row + 1, row, header_row)
        row = max(row, chart_bottom)

    closing_soon = [
        tender
        for tender in sort_for_review(actionable)
        if (days := _days_left(tender.deadline, now)) is not None and 0 <= days <= 3
    ][:10]
    if closing_soon:
        row += 2
        row = _dashboard_section(sheet, row, "Скоро закрываются")
        soon_headers = ["Название", "Приоритет", "Сумма", "Срок подачи", "Дней"]
        for index, header in enumerate(soon_headers, start=1):
            sheet.cell(row=row, column=index, value=header)
        _style_header(sheet, len(soon_headers), row=row)
        for tender in closing_soon:
            row += 1
            title_cell = sheet.cell(row=row, column=1, value=_safe(tender.title))
            title_cell.alignment = WRAP_TOP
            if tender.url:
                title_cell.hyperlink = tender.url
                title_cell.font = LINK_FONT
            sheet.cell(
                row=row,
                column=2,
                value=PRIORITY_LABELS.get(tender.review_priority or "", tender.review_priority or ""),
            )
            price_cell = sheet.cell(row=row, column=3, value=tender.price)
            price_cell.number_format = PRICE_FORMAT
            deadline_cell = sheet.cell(row=row, column=4, value=tender.deadline)
            deadline_cell.number_format = DATE_FORMAT
            days_cell = sheet.cell(row=row, column=5, value=_days_left(tender.deadline, now))
            days_cell.fill = PatternFill("solid", start_color=DEADLINE_URGENT_FILL)
            _border_row(sheet, row, 5)
            sheet.row_dimensions[row].height = ROW_LINE_HEIGHT * max(1, min(ceil(len(tender.title) / 53), 4)) + 3

    if excluded:
        row += 2
        row = _dashboard_section(sheet, row, "Топ причин отсева")
        sheet.cell(row=row, column=1, value="Причина")
        sheet.cell(row=row, column=2, value="Карточек")
        _style_header(sheet, 2, row=row)
        reasons = _count_by(excluded, lambda tender: tender.exclude_reason or "без причины")
        for name, count in reasons[:8]:
            row += 1
            reason_cell = sheet.cell(row=row, column=1, value=_safe(name))
            reason_cell.alignment = WRAP_TOP
            sheet.cell(row=row, column=2, value=count)
            _border_row(sheet, row, 2)

    row += 2
    row = _dashboard_section(sheet, row, "Топ горячих")
    top_headers = ["Название", "Регион", "Сумма", "Срок подачи", "Дней до срока"]
    for index, header in enumerate(top_headers, start=1):
        sheet.cell(row=row, column=index, value=header)
    _style_header(sheet, len(top_headers), row=row)
    for tender in sort_for_review(hot)[:10]:
        row += 1
        title_cell = sheet.cell(row=row, column=1, value=_safe(tender.title))
        title_cell.alignment = WRAP_TOP
        if tender.url:
            title_cell.hyperlink = tender.url
            title_cell.font = LINK_FONT
        sheet.cell(row=row, column=2, value=tender.region or "")
        price_cell = sheet.cell(row=row, column=3, value=tender.price)
        price_cell.number_format = PRICE_FORMAT
        deadline_cell = sheet.cell(row=row, column=4, value=tender.deadline)
        deadline_cell.number_format = DATE_FORMAT
        days = _days_left(tender.deadline, now)
        sheet.cell(row=row, column=5, value=days if days is not None else "")
        _border_row(sheet, row, 5)
        sheet.row_dimensions[row].height = ROW_LINE_HEIGHT * max(1, min(ceil(len(tender.title) / 53), 4)) + 3

    if source_health:
        row += 2
        row = _dashboard_section(sheet, row, "Здоровье источников")
        health_headers = ["Источник", "Статус", "Найдено", "Время, сек", "Детали"]
        for index, header in enumerate(health_headers, start=1):
            sheet.cell(row=row, column=index, value=header)
        _style_header(sheet, len(health_headers), row=row)
        for item in source_health:
            row += 1
            sheet.cell(row=row, column=1, value=item.source)
            status_cell = sheet.cell(row=row, column=2, value=item.status)
            status_cell.fill = PatternFill(
                "solid", start_color=HEALTH_FILLS.get(item.status, HEALTH_BAD_FILL)
            )
            sheet.cell(row=row, column=3, value=item.found)
            sheet.cell(row=row, column=4, value=item.elapsed_seconds)
            detail_cell = sheet.cell(row=row, column=5, value=_safe(item.detail))
            detail_cell.alignment = WRAP_TOP
            _border_row(sheet, row, 5)


def export_excel(
    hot: list[TenderRecord],
    review: list[TenderRecord],
    wide: list[TenderRecord],
    excluded: list[TenderRecord],
    output_path: Path,
    *,
    new_tenders: list[TenderRecord] | None = None,
    now: datetime | None = None,
    source_health: list[SourceHealth] | None = None,
    manual_selections: Mapping[SelectionKey, SelectionValue] | None = None,
    new_keys: set[str] | None = None,
) -> Path:
    current = now or datetime.now()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    dashboard = workbook.active
    dashboard.title = "Дашборд"
    _build_dashboard(
        dashboard,
        hot=hot,
        review=review,
        wide=wide,
        excluded=excluded,
        new_count=len(new_tenders) if new_tenders is not None else None,
        now=current,
        source_health=source_health,
    )

    picked = [
        tender
        for tender in hot + review + wide + excluded
        if _selection_for(tender, manual_selections)[0].startswith(PICKED_CHOICES)
    ]
    if picked:
        _append_rows(
            workbook.create_sheet("Мой отбор"), sort_for_review(picked), current, manual_selections, new_keys
        )

    if new_tenders is not None:
        _append_rows(
            workbook.create_sheet("Новые"), sort_for_review(new_tenders), current, manual_selections, new_keys
        )
    _append_rows(workbook.create_sheet("Горячие"), sort_for_review(hot), current, manual_selections, new_keys)
    _append_rows(
        workbook.create_sheet("На проверку"), sort_for_review(review), current, manual_selections, new_keys
    )
    _append_rows(
        workbook.create_sheet("Широкий хвост"), sort_for_review(wide), current, manual_selections, new_keys
    )
    _append_rows(
        workbook.create_sheet("Отсеянные"), sort_for_review(excluded), current, manual_selections, new_keys
    )

    for name, color in TAB_COLORS.items():
        if name in workbook.sheetnames:
            workbook[name].sheet_properties.tabColor = color

    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        # Файл открыт в Excel — сохраняем рядом с меткой времени, чтобы прогон не падал.
        fallback = output_path.with_name(
            f"{output_path.stem}_{datetime.now():%H%M%S}{output_path.suffix}"
        )
        try:
            workbook.save(fallback)
        except PermissionError:
            fallback = output_path.with_name(f"{output_path.stem}_{uuid4().hex[:8]}{output_path.suffix}")
            workbook.save(fallback)
        print(f"Файл {output_path.name} занят (открыт в Excel) — отчет сохранен как {fallback.name}")
        return fallback
