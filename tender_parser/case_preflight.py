from __future__ import annotations

import csv
from difflib import SequenceMatcher
import json
import re
import shutil
import subprocess
import tempfile
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import Literal
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook
from pypdf import PdfReader

from tender_parser.documents import SUPPORTED_SUFFIXES


Severity = Literal["blocker", "risk", "action", "info"]


@dataclass(frozen=True)
class TextSegment:
    source: str
    locator: str
    text: str


@dataclass(frozen=True)
class DocumentRecord:
    path: str
    document_type: str
    searchable: bool
    text_characters: int
    pages_or_sheets: int | None = None
    note: str = ""


@dataclass(frozen=True)
class Finding:
    severity: Severity
    code: str
    message: str
    source: str = ""
    locator: str = ""
    excerpt: str = ""


@dataclass(frozen=True)
class ItemCandidate:
    line_id: str
    name: str
    quantity: str
    unit: str
    required_specs: str
    source: str
    locator: str
    confidence: str


@dataclass(frozen=True)
class PreflightResult:
    case_id: str
    generated_at: str
    documents: list[DocumentRecord] = field(default_factory=list)
    document_types_found: list[str] = field(default_factory=list)
    metadata_candidates: dict[str, list[dict[str, str]]] = field(default_factory=dict)
    item_candidates: list[ItemCandidate] = field(default_factory=list)
    findings: list[Finding] = field(default_factory=list)
    questions: list[str] = field(default_factory=list)
    ready_for_product_search: bool = False


DOCUMENT_TYPES = {
    "technical_specification": (
        "техническое задание",
        "описание объекта закупки",
        "технические характеристики",
        "спецификация",
    ),
    "contract_draft": ("проект контракта", "проект договора", "условия контракта"),
    "notice": ("извещение об осуществлении закупки", "извещение о закупке", "информационная карта"),
    "application_requirements": (
        "требования к содержанию заявки",
        "требования к заявке",
        "инструкция по заполнению заявки",
        "инструкция по ее заполнению",
        "требования к участникам закупки",
    ),
    "price_justification": ("обоснование начальной", "обоснование нмцк", "расчет нмцк"),
}

DOCUMENT_TYPE_LABELS = {
    "technical_specification": "Техническое задание / описание объекта закупки",
    "contract_draft": "Проект контракта / договора",
    "notice": "Извещение / информационная карта",
    "application_requirements": "Требования к заявке и участнику",
    "price_justification": "Обоснование НМЦК",
    "other": "Прочий документ",
    "archive": "Архив документации",
    "unreadable": "Не прочитан",
}

ARCHIVE_SUFFIXES = {".rar", ".zip"}
LEGACY_DOCUMENT_SUFFIXES = {".doc"}

UNIT_PATTERN = r"(?:шт\.?|штук\w*|компл\.?|комплект\w*|ед\.?|упак\.?|упаков\w*|парт\w*)"
MONEY_PATTERN = r"(\d{1,3}(?:[\s\u00a0]\d{3})*(?:[.,]\d{1,2})?|\d+(?:[.,]\d{1,2})?)"


def analyze_case_documents(case_dir: Path) -> PreflightResult:
    case_payload = _load_case_payload(case_dir)
    documents_dir = case_dir / "documents"
    records: list[DocumentRecord] = []
    segments: list[TextSegment] = []
    item_candidates: list[ItemCandidate] = []
    findings: list[Finding] = []

    source_files = [path for path in sorted(documents_dir.rglob("*")) if path.is_file()] if documents_dir.exists() else []
    if not source_files:
        findings.append(Finding("blocker", "documents_missing", "Документация закупки не загружена."))

    with tempfile.TemporaryDirectory(prefix="tender-preflight-") as temporary:
        files: list[tuple[Path, str]] = []
        for archive_number, path in enumerate(source_files, start=1):
            relative = str(path.relative_to(documents_dir))
            if path.suffix.lower() not in ARCHIVE_SUFFIXES:
                files.append((path, relative))
                continue
            extract_dir = Path(temporary) / f"archive-{archive_number}"
            try:
                extracted = _extract_archive(path, extract_dir)
            except (OSError, ValueError, subprocess.SubprocessError) as exc:
                records.append(DocumentRecord(relative, "unreadable", False, 0, note=exc.__class__.__name__))
                findings.append(Finding("blocker", "archive_unreadable", "Архив документации не удалось распаковать.", relative))
                continue
            records.append(
                DocumentRecord(relative, "archive", False, 0, note=f"Распаковано файлов: {len(extracted)}")
            )
            files.extend((item, f"{relative}::{item.relative_to(extract_dir)}") for item in extracted)

        for path, relative in files:
            supported = path.suffix.lower() in SUPPORTED_SUFFIXES or path.suffix.lower() in LEGACY_DOCUMENT_SUFFIXES
            if not supported:
                records.append(DocumentRecord(relative, "other", False, 0, note="Формат не поддерживается для анализа"))
                findings.append(
                    Finding("action", "unsupported_document", "Файл нужно преобразовать в PDF, DOCX или XLSX.", relative)
                )
                continue
            try:
                file_segments, page_count = _extract_segments(path, relative)
            except Exception as exc:
                records.append(DocumentRecord(relative, "unreadable", False, 0, note=exc.__class__.__name__))
                findings.append(Finding("blocker", "document_unreadable", "Документ не удалось прочитать.", relative))
                continue
            text = " ".join(segment.text for segment in file_segments).strip()
            document_type = _classify_document(relative, text)
            records.append(
                DocumentRecord(
                    relative,
                    document_type,
                    bool(text),
                    len(text),
                    pages_or_sheets=page_count,
                    note="Нужен OCR" if path.suffix.lower() == ".pdf" and not text else "",
                )
            )
            if path.suffix.lower() == ".pdf" and not text:
                findings.append(Finding("blocker", "ocr_required", "PDF не содержит текстового слоя; требуется OCR.", relative))
            segments.extend(file_segments)
            if path.suffix.lower() == ".xlsx" or document_type in {"technical_specification", "price_justification"}:
                item_candidates.extend(_extract_item_candidates(path, relative, file_segments))

    document_types = list(dict.fromkeys(record.document_type for record in records if record.document_type != "other"))
    _check_document_completeness(document_types, findings, str(case_payload.get("law") or ""))
    metadata = _extract_metadata(segments)
    _collect_content_findings(segments, findings, case_payload)
    item_candidates = _reconcile_items(item_candidates, findings)

    if not item_candidates:
        findings.append(
            Finding(
                "action",
                "items_not_extracted",
                "Позиции автоматически не выделены. Заполните items.csv вручную по ТЗ или спецификации.",
            )
        )

    questions = _build_questions(findings, metadata)
    blockers = [finding for finding in findings if finding.severity == "blocker"]
    ready_for_product_search = not blockers and bool(item_candidates or _existing_items(case_dir))
    return PreflightResult(
        case_id=str(case_payload.get("case_id") or case_dir.name),
        generated_at=datetime.now().astimezone().isoformat(timespec="seconds"),
        documents=records,
        document_types_found=document_types,
        metadata_candidates=metadata,
        item_candidates=item_candidates,
        findings=_deduplicate_findings(findings),
        questions=questions,
        ready_for_product_search=ready_for_product_search,
    )


def export_preflight(result: PreflightResult, output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "preflight.json"
    markdown_path = output_dir / "preflight.md"
    items_path = output_dir / "items_draft.csv"
    questions_path = output_dir / "customer_questions.txt"

    json_path.write_text(json.dumps(asdict(result), ensure_ascii=False, indent=2), encoding="utf-8")
    markdown_path.write_text(_render_markdown(result), encoding="utf-8")
    questions_path.write_text(
        "\n".join(f"{index}. {question}" for index, question in enumerate(result.questions, start=1))
        or "Вопросы автоматически не сформированы.\n",
        encoding="utf-8-sig",
    )
    with items_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow(
            [
                "line_id",
                "name",
                "quantity",
                "unit",
                "required_specs",
                "mandatory",
                "source",
                "locator",
                "confidence",
                "owner_status",
            ]
        )
        for item in result.item_candidates:
            writer.writerow(
                [
                    item.line_id,
                    item.name,
                    item.quantity,
                    item.unit,
                    item.required_specs,
                    "yes",
                    item.source,
                    item.locator,
                    item.confidence,
                    "Проверить",
                ]
            )
    return {"json": json_path, "markdown": markdown_path, "items": items_path, "questions": questions_path}


def _load_case_payload(case_dir: Path) -> dict[str, object]:
    path = case_dir / "case.json"
    if not path.exists():
        raise FileNotFoundError(f"Не найден паспорт тендерного дела: {path}")
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("case.json должен содержать объект")
    return payload


def _extract_segments(path: Path, relative: str) -> tuple[list[TextSegment], int | None]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        reader = PdfReader(path)
        segments: list[TextSegment] = []
        for index, page in enumerate(reader.pages, start=1):
            values = [page.extract_text() or ""]
            for annotation in page.annotations or []:
                value = annotation.get_object().get("/Contents")
                if value:
                    values.append(str(value))
            segments.append(TextSegment(relative, f"стр. {index}", _clean_text(" ".join(values))))
        return segments, len(reader.pages)
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            result = []
            for sheet in workbook.worksheets:
                text = "\n".join(
                    " | ".join(str(cell) for cell in row if cell is not None)
                    for row in sheet.iter_rows(values_only=True)
                )
                result.append(TextSegment(relative, f"лист «{sheet.title}»", _clean_text(text)))
            return result, len(workbook.worksheets)
        finally:
            workbook.close()
    if suffix == ".docx":
        with ZipFile(path) as archive:
            root = ElementTree.fromstring(archive.read("word/document.xml"))
        text = " ".join(element.text or "" for element in root.iter())
        segments = [TextSegment(relative, "", _clean_text(text))]
        namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
        for table_number, table in enumerate(root.iter(f"{namespace}tbl"), start=1):
            for row_number, row in enumerate(table.findall(f"{namespace}tr"), start=1):
                cells = []
                for cell in row.findall(f"{namespace}tc"):
                    value = " ".join(element.text or "" for element in cell.iter(f"{namespace}t"))
                    cells.append(_clean_text(value))
                if any(cells):
                    segments.append(
                        TextSegment(relative, f"таблица {table_number}, строка {row_number}", " | ".join(cells))
                    )
        return segments, None
    if suffix == ".doc":
        return _extract_legacy_doc(path, relative)
    if suffix == ".json":
        text = json.dumps(json.loads(_read_text(path)), ensure_ascii=False)
    elif suffix in {".html", ".htm"}:
        text = re.sub(r"<[^>]+>", " ", _read_text(path))
    else:
        text = _read_text(path)
    return [TextSegment(relative, "", _clean_text(text))], None


def _extract_archive(path: Path, destination: Path) -> list[Path]:
    destination.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() == ".zip":
        with ZipFile(path) as archive:
            members = [item for item in archive.infolist() if not item.is_dir()]
            for member in members:
                target = _safe_archive_target(destination, member.filename)
                target.parent.mkdir(parents=True, exist_ok=True)
                with archive.open(member) as source, target.open("wb") as output:
                    shutil.copyfileobj(source, output)
    else:
        command = shutil.which("tar")
        if not command:
            raise OSError("Не найден системный распаковщик tar для RAR")
        listing = subprocess.run(
            [command, "-tf", str(path)],
            check=True,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        for member in listing.stdout.splitlines():
            if member.strip():
                _safe_archive_target(destination, member)
        subprocess.run([command, "-xf", str(path), "-C", str(destination)], check=True, capture_output=True)
    return [item for item in sorted(destination.rglob("*")) if item.is_file() and not item.is_symlink()]


def _safe_archive_target(destination: Path, member_name: str) -> Path:
    normalized = member_name.replace("\\", "/")
    member = PurePosixPath(normalized)
    if member.is_absolute() or ".." in member.parts or (member.parts and ":" in member.parts[0]):
        raise ValueError(f"Небезопасный путь в архиве: {member_name}")
    target = destination.joinpath(*member.parts)
    try:
        target.resolve().relative_to(destination.resolve())
    except ValueError as exc:
        raise ValueError(f"Небезопасный путь в архиве: {member_name}") from exc
    return target


def _extract_legacy_doc(path: Path, relative: str) -> tuple[list[TextSegment], int | None]:
    if not _word_automation_available():
        raise OSError("Для чтения DOC требуется установленный Microsoft Word")
    with tempfile.TemporaryDirectory(prefix="tender-doc-") as temporary:
        converted = Path(temporary) / "converted.docx"
        import win32com.client  # type: ignore[import-untyped]

        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        word.AutomationSecurity = 3
        document = None
        try:
            document = word.Documents.Open(
                str(path.resolve()),
                ConfirmConversions=False,
                ReadOnly=True,
                AddToRecentFiles=False,
            )
            document.SaveAs2(str(converted), FileFormat=16)
            document.Close(False)
            document = None
        finally:
            if document is not None:
                document.Close(False)
            word.Quit()
        return _extract_segments(converted, relative)


def _word_automation_available() -> bool:
    try:
        import win32com.client  # type: ignore[import-untyped]  # noqa: F401
    except ImportError:
        return False
    return True


def _classify_document(filename: str, text: str) -> str:
    normalized_filename = filename.lower().replace("ё", "е")
    filename_scores = {
        kind: sum(1 for marker in markers if marker in normalized_filename)
        for kind, markers in DOCUMENT_TYPES.items()
    }
    filename_best = max(filename_scores, key=filename_scores.get)
    if filename_scores[filename_best]:
        return filename_best
    sample = f"{filename} {text[:30000]}".lower().replace("ё", "е")
    scores = {kind: sum(1 for marker in markers if marker in sample) for kind, markers in DOCUMENT_TYPES.items()}
    best = max(scores, key=scores.get)
    return best if scores[best] else "other"


def _extract_metadata(segments: list[TextSegment]) -> dict[str, list[dict[str, str]]]:
    patterns = {
        "nmck": (
            rf"(?:нмцк|начальн\w*\s+\(максимальн\w*\)\s+цен\w*|начальн\w*\s+максимальн\w*\s+цен\w*)[^\d]{{0,80}}{MONEY_PATTERN}\s*(?:руб|₽)",
        ),
        "payment_days": (
            r"(?:оплат\w*|срок\s+оплаты)[^.\n]{0,160}?(?:в\s+течение\s+)?(\d{1,3})\s*(?:рабоч\w*|календарн\w*)?\s*дн",
        ),
        "delivery_days": (
            r"(?:срок\w*\s+поставк\w*|поставк\w*\s+товар\w*)[^.\n]{0,180}?(?:в\s+течение\s+)?(\d{1,3})\s*(?:рабоч\w*|календарн\w*)?\s*дн",
        ),
        "application_security": (
            rf"обеспечени\w*\s+заявк\w*[^.\n]{{0,120}}?{MONEY_PATTERN}\s*(?:руб|₽|%)",
        ),
        "contract_security": (
            rf"обеспечени\w*\s+(?:исполнени\w*\s+)?контракт\w*[^.\n]{{0,120}}?{MONEY_PATTERN}\s*(?:руб|₽|%)",
        ),
    }
    result: dict[str, list[dict[str, str]]] = {}
    for key, expressions in patterns.items():
        values = []
        for segment in segments:
            for expression in expressions:
                for match in re.finditer(expression, segment.text, flags=re.IGNORECASE):
                    values.append(
                        {
                            "value": match.group(1).replace("\u00a0", " ").strip(),
                            "source": segment.source,
                            "locator": segment.locator,
                            "excerpt": _excerpt(segment.text, match.start(), match.end()),
                        }
                    )
        if values:
            result[key] = _unique_dicts(values)
    return result


def _extract_item_candidates(path: Path, relative: str, segments: list[TextSegment]) -> list[ItemCandidate]:
    if path.suffix.lower() == ".xlsx":
        return _extract_xlsx_items(path, relative)
    if path.suffix.lower() in {".doc", ".docx"}:
        table_candidates = _extract_word_table_items(segments, relative)
        if table_candidates:
            return table_candidates
    candidates: list[ItemCandidate] = []
    pattern = re.compile(
        rf"(?m)^\s*(\d{{1,4}})[.)]\s+(.{{3,180}}?)\s+(\d+(?:[.,]\d+)?)\s*({UNIT_PATTERN})\s*$",
        re.IGNORECASE,
    )
    for segment in segments:
        for match in pattern.finditer(segment.text):
            candidates.append(
                ItemCandidate(
                    line_id=match.group(1),
                    name=_clean_text(match.group(2)),
                    quantity=match.group(3).replace(",", "."),
                    unit=match.group(4),
                    required_specs="",
                    source=relative,
                    locator=segment.locator,
                    confidence="medium",
                )
            )
    return candidates


def _extract_word_table_items(segments: list[TextSegment], relative: str) -> list[ItemCandidate]:
    candidates: list[ItemCandidate] = []
    current: dict[str, object] | None = None
    valid_tables = {
        segment.locator.split(", строка", maxsplit=1)[0]
        for segment in segments
        if segment.locator.startswith("таблица ")
        and "наименование" in segment.text.lower()
        and any(marker in segment.text.lower() for marker in ("кол-во", "количество", "кол во"))
    }

    def finish() -> None:
        nonlocal current
        if current is None:
            return
        specs = current.pop("specs")
        candidates.append(
            ItemCandidate(
                required_specs="; ".join(specs),
                confidence="high",
                **current,
            )
        )
        current = None

    for segment in segments:
        table_name = segment.locator.split(", строка", maxsplit=1)[0]
        if table_name not in valid_tables:
            continue
        cells = [_clean_text(value) for value in segment.text.split("|")]
        if len(cells) < 5:
            continue
        line_value = cells[0].rstrip(". ")
        quantity = cells[-1].replace(",", ".")
        starts_item = bool(
            re.fullmatch(r"\d{1,4}", line_value)
            and cells[1]
            and not re.fullmatch(r"\d+(?:\.\d+)?", cells[1])
            and re.fullmatch(r"\d+(?:\.\d+)?", quantity)
        )
        if starts_item:
            finish()
            current = {
                "line_id": line_value,
                "name": cells[1].replace("/", " "),
                "quantity": quantity,
                "unit": cells[-2] or "шт.",
                "source": relative,
                "locator": segment.locator,
                "specs": [],
            }
        if current is not None and len(cells) >= 5 and cells[3]:
            characteristic = cells[3]
            value = cells[4]
            current["specs"].append(f"{characteristic}: {value}" if value else characteristic)  # type: ignore[union-attr]
    finish()
    return candidates


def _extract_xlsx_items(path: Path, relative: str) -> list[ItemCandidate]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    candidates: list[ItemCandidate] = []
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            header_index, columns = _find_item_header(rows)
            if header_index is None or "name" not in columns or "quantity" not in columns:
                continue
            for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                name = _cell(row, columns.get("name"))
                quantity = _cell(row, columns.get("quantity"))
                if not name or not quantity or not re.search(r"\d", quantity):
                    continue
                line_id = _cell(row, columns.get("line_id")) or str(len(candidates) + 1)
                candidates.append(
                    ItemCandidate(
                        line_id=line_id,
                        name=name,
                        quantity=quantity.replace(",", "."),
                        unit=_cell(row, columns.get("unit")) or "шт.",
                        required_specs=_cell(row, columns.get("specs")),
                        source=relative,
                        locator=f"лист «{sheet.title}», строка {row_number}",
                        confidence="high",
                    )
                )
    finally:
        workbook.close()
    return candidates


def _find_item_header(rows: list[list[object]], limit: int = 30) -> tuple[int | None, dict[str, int]]:
    aliases = {
        "line_id": ("№", "номер", "п/п", "позиция"),
        "name": ("наименование", "наименование товара", "товар", "предмет закупки"),
        "quantity": ("количество", "кол-во", "объем"),
        "unit": ("единица измерения", "ед. изм.", "единица"),
        "specs": ("характеристики", "технические характеристики", "описание", "требования"),
    }
    for row_index, row in enumerate(rows[:limit]):
        normalized = [_normalize_header(cell) for cell in row]
        columns: dict[str, int] = {}
        for key, options in aliases.items():
            for index, value in enumerate(normalized):
                if _matches_header(value, options):
                    columns[key] = index
                    break
        if (
            "name" in columns
            and "quantity" in columns
            and columns["name"] != columns["quantity"]
            and len(set(columns.values())) == len(columns)
        ):
            return row_index, columns
    return None, {}


def _matches_header(value: str, options: tuple[str, ...]) -> bool:
    if not value or len(value) > 120:
        return False
    return any(value == option or value.startswith(f"{option} ") or value.startswith(f"{option},") for option in options)


def _check_document_completeness(types: list[str], findings: list[Finding], law: str) -> None:
    required = {
        "technical_specification": "Не найдено техническое задание или описание объекта закупки.",
        "contract_draft": "Не найден проект контракта/договора.",
        "notice": "Не найдено извещение или информационная карта.",
    }
    for kind, message in required.items():
        if kind not in types:
            findings.append(Finding("blocker", f"missing_{kind}", message))
    if law.lower().startswith("223") and "application_requirements" not in types:
        findings.append(
            Finding("action", "check_223_rules", "Для закупки по 223-ФЗ проверьте положение о закупке заказчика и требования к заявке.")
        )


def _collect_content_findings(
    segments: list[TextSegment], findings: list[Finding], case_payload: dict[str, object]
) -> None:
    checks = [
        ("national_regime", "action", ("постановлени", "1875"), "Проверить запрет, ограничение или преимущество по ПП РФ № 1875 и документы о стране происхождения."),
        ("certificates", "action", ("сертификат",), "Проверить обязательные сертификаты/декларации и возможность получить их до подачи."),
        ("installation", "risk", ("монтаж",), "В документации упомянут монтаж; нужен отдельный расчет состава и стоимости работ."),
        ("license", "risk", ("лицензи",), "В документации упомянута лицензия; проверить применимость требования к предмету закупки и участнику."),
        ("sro", "risk", ("сро",), "В документации упомянуто СРО; проверить применимость требования к поставке товара."),
        ("concrete_values", "action", ("конкретн", "показател"), "Нужно подготовить таблицу конкретных показателей без двусмысленных значений."),
        ("guarantee", "action", ("гарантийн",), "Проверить гарантийный срок, место и порядок гарантийного обслуживания."),
    ]
    for segment in segments:
        lowered = segment.text.lower().replace("ё", "е")
        for code, severity, tokens, message in checks:
            if all(token in lowered for token in tokens):
                start = min(lowered.find(token) for token in tokens if token in lowered)
                findings.append(
                    Finding(severity, code, message, segment.source, segment.locator, _excerpt(segment.text, start, start + 180))
                )
    combined = " ".join(segment.text.lower().replace("ё", "е") for segment in segments)
    excluded = {
        "road_construction": ("строительство автомобильной дороги", "ремонт автомобильной дороги"),
        "capital_construction": ("объект капитального строительства",),
        "fuel": ("бензин", "дизельное топливо", "горюче-смазочн"),
        "medicines": ("лекарственный препарат", "лекарственные средства"),
    }
    for code, markers in excluded.items():
        if any(marker in combined for marker in markers):
            findings.append(
                Finding("risk", f"excluded_scope_{code}", "Обнаружена стоп-тема. Подтвердите, что закупка относится к допустимому товарному направлению.")
            )
    if bool(case_payload.get("requires_installation")) and not any(f.code == "installation" for f in findings):
        findings.append(Finding("risk", "installation", "В паспорте дела указан монтаж; требуется отдельный расчет."))


def _build_questions(findings: list[Finding], metadata: dict[str, list[dict[str, str]]]) -> list[str]:
    questions = []
    for finding in findings:
        if finding.code.startswith("missing_"):
            questions.append(f"Получить у заказчика/из ЕИС документ: {finding.message}")
        elif finding.code == "ocr_required":
            questions.append(f"Получить читаемую версию или выполнить OCR файла «{finding.source}».")
        elif finding.code == "national_regime":
            questions.append("Уточнить применяемую защитную меру и допустимое подтверждение страны происхождения каждой позиции.")
        elif finding.code == "certificates":
            questions.append("Уточнить перечень обязательных сертификатов/деклараций и момент их предоставления.")
        elif finding.code == "installation":
            questions.append("Уточнить точный состав, адрес, сроки и условия приемки монтажных работ.")
        elif finding.code in {"license", "sro"}:
            questions.append("Запросить разъяснение о применимости специального разрешительного требования к поставщику товара.")
    if "payment_days" not in metadata:
        questions.append("Подтвердить срок и условие оплаты после приемки товара.")
    if "delivery_days" not in metadata:
        questions.append("Подтвердить срок поставки и дату, от которой он исчисляется.")
    return list(dict.fromkeys(questions))


def _render_markdown(result: PreflightResult) -> str:
    readiness = "ДА" if result.ready_for_product_search else "НЕТ"
    lines = [
        f"# Предварительный анализ дела {result.case_id}",
        "",
        "> Автоматический черновик. Перед решением об участии выводы и позиции проверяет владелец.",
        "",
        f"Готово к поиску товаров: **{readiness}**",
        "",
        "## Документы",
        "",
    ]
    if not result.documents:
        lines.append("- Документы не загружены.")
    for document in result.documents:
        label = DOCUMENT_TYPE_LABELS.get(document.document_type, document.document_type)
        status = "читается" if document.searchable else "не читается"
        lines.append(f"- `{document.path}` — {label}; {status}; символов: {document.text_characters}.")
    lines.extend(["", "## Выводы и действия", ""])
    for finding in result.findings:
        evidence = ""
        if finding.source:
            evidence = f" — `{finding.source}` {finding.locator}".rstrip()
        lines.append(f"- **{finding.severity.upper()}**: {finding.message}{evidence}")
    lines.extend(["", "## Кандидаты позиций", ""])
    if not result.item_candidates:
        lines.append("Позиции автоматически не выделены.")
    else:
        lines.append("| № | Наименование | Количество | Источник | Доверие |")
        lines.append("|---|---|---:|---|---|")
        for item in result.item_candidates:
            name = item.name.replace("|", "/")
            lines.append(f"| {item.line_id} | {name} | {item.quantity} {item.unit} | {item.source}, {item.locator} | {item.confidence} |")
    lines.extend(["", "## Вопросы", ""])
    for index, question in enumerate(result.questions, start=1):
        lines.append(f"{index}. {question}")
    lines.append("")
    return "\n".join(lines)


def _existing_items(case_dir: Path) -> bool:
    path = case_dir / "items.csv"
    if not path.exists():
        return False
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return any((row.get("name") or "").strip() for row in csv.DictReader(handle, delimiter=";"))


def _reconcile_items(items: list[ItemCandidate], findings: list[Finding]) -> list[ItemCandidate]:
    result: list[ItemCandidate] = []
    for item in items:
        match_index = next(
            (
                index
                for index, existing in enumerate(result)
                if existing.line_id == item.line_id
                and existing.quantity == item.quantity
            ),
            None,
        )
        if match_index is None:
            if not any(
                _canonical_item_name(existing.name) == _canonical_item_name(item.name)
                and existing.quantity == item.quantity
                and existing.source == item.source
                for existing in result
            ):
                result.append(item)
            continue
        existing = result[match_index]
        if _item_names_materially_differ(existing.name, item.name):
            findings.append(
                Finding(
                    "risk",
                    "item_document_conflict",
                    f"По позиции {item.line_id} наименования или коды в документах различаются: «{existing.name}» / «{item.name}».",
                    item.source,
                    item.locator,
                )
            )
        if item.required_specs and not existing.required_specs:
            result[match_index] = item
    return result


def _canonical_item_name(value: str) -> str:
    return re.sub(r"[^a-zа-яё0-9]+", " ", value.lower()).strip()


def _item_base_name(value: str) -> str:
    without_code = re.sub(r"\b\d{2}(?:\.\d{2}){2}\.\d{3}\b", " ", value)
    return _canonical_item_name(without_code)


def _item_names_materially_differ(left: str, right: str) -> bool:
    code_pattern = r"\b\d{2}(?:\.\d{2}){2}\.\d{3}\b"
    left_codes = set(re.findall(code_pattern, left))
    right_codes = set(re.findall(code_pattern, right))
    if left_codes and right_codes and left_codes != right_codes:
        return True
    return SequenceMatcher(None, _item_base_name(left), _item_base_name(right)).ratio() < 0.8


def _deduplicate_findings(findings: list[Finding]) -> list[Finding]:
    result: list[Finding] = []
    seen: set[tuple[str, str]] = set()
    for finding in findings:
        key = (finding.code, finding.source)
        if key not in seen:
            seen.add(key)
            result.append(finding)
    return result


def _unique_dicts(values: list[dict[str, str]]) -> list[dict[str, str]]:
    result = []
    seen: set[tuple[str, str, str]] = set()
    for value in values:
        key = (value["value"], value["source"], value["locator"])
        if key not in seen:
            seen.add(key)
            result.append(value)
    return result


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("ё", "е"))


def _cell(row: list[object], index: int | None) -> str:
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return _clean_text(str(row[index]))


def _clean_text(value: str) -> str:
    return re.sub(r"[ \t\r\f\v]+", " ", value).strip()


def _excerpt(text: str, start: int, end: int, radius: int = 100) -> str:
    return _clean_text(text[max(0, start - radius) : min(len(text), end + radius)])[:500]


def _read_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")
