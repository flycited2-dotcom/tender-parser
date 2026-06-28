from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path
from time import monotonic
from xml.etree import ElementTree

from openpyxl import load_workbook

from tender_parser.models import TenderRecord
from tender_parser.run_report import SourceFetchResult, SourceHealth


HEADER_ALIASES = {
    "title": {"title", "name", "название", "наименование", "предмет", "предмет закупки"},
    "url": {"url", "link", "ссылка", "адрес"},
    "source": {"source", "источник", "площадка", "этп"},
    "number": {"number", "номер", "номер закупки", "номер процедуры", "извещение"},
    "customer": {"customer", "заказчик", "организатор"},
    "region": {"region", "регион", "место поставки", "адрес поставки"},
    "price": {"price", "сумма", "нмцк", "начальная цена", "цена"},
    "deadline": {"deadline", "срок подачи", "дата окончания", "окончание подачи"},
    "published_at": {"published_at", "опубликовано", "дата публикации", "размещено"},
    "raw_text": {"raw_text", "описание", "текст"},
}


class ImportFolderSource:
    def __init__(self, imports_dir: Path) -> None:
        self.imports_dir = imports_dir

    def fetch_keywords(self, keywords: list[str]) -> list[TenderRecord]:
        return self.fetch_with_report(keywords).tenders

    def fetch_with_report(self, keywords: list[str]) -> SourceFetchResult:
        started_at = monotonic()
        if not self.imports_dir.exists():
            return SourceFetchResult(
                health=[
                    SourceHealth(
                        source="ImportFolderSource",
                        status="empty",
                        found=0,
                        elapsed_seconds=round(monotonic() - started_at, 3),
                        detail=f"folder missing: {self.imports_dir}",
                    )
                ]
            )

        tenders: list[TenderRecord] = []
        errors: list[str] = []
        for path in sorted(self.imports_dir.iterdir()):
            if not path.is_file():
                continue
            try:
                tenders.extend(_read_file(path))
            except (OSError, ValueError, ElementTree.ParseError) as exc:
                errors.append(f"{path.name}: {exc}")

        status = "ok" if tenders else "empty"
        if errors and tenders:
            status = "partial"
        elif errors:
            status = "error"
        return SourceFetchResult(
            tenders=tenders,
            health=[
                SourceHealth(
                    source="ImportFolderSource",
                    status=status,
                    found=len(tenders),
                    elapsed_seconds=round(monotonic() - started_at, 3),
                    detail="; ".join(errors),
                )
            ],
            errors=errors,
        )


def _read_file(path: Path) -> list[TenderRecord]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix == ".xml":
        return _read_xml(path)
    return []


def _read_csv(path: Path) -> list[TenderRecord]:
    text = path.read_text(encoding="utf-8-sig")
    sample = text[:2048]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    rows = csv.DictReader(text.splitlines(), delimiter=delimiter)
    return [_record_from_row(_normalize_row(row), path) for row in rows if _normalize_row(row)]


def _read_xlsx(path: Path) -> list[TenderRecord]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    result: list[TenderRecord] = []
    for row in rows[1:]:
        values = {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        normalized = _normalize_row(values)
        if normalized:
            result.append(_record_from_row(normalized, path))
    return result


def _read_xml(path: Path) -> list[TenderRecord]:
    root = ElementTree.fromstring(path.read_text(encoding="utf-8"))
    result: list[TenderRecord] = []
    for element in root.iter():
        if element is root or element.tag.lower() not in {"tender", "item", "procedure", "purchase"}:
            continue
        row = {child.tag: child.text or "" for child in list(element)}
        normalized = _normalize_row(row)
        if normalized:
            result.append(_record_from_row(normalized, path))
    return result


def _normalize_row(row: dict[object, object]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in row.items():
        canonical = _canonical_header(str(key or ""))
        if canonical and value is not None:
            normalized[canonical] = str(value).strip()
    return normalized


def _canonical_header(value: str) -> str | None:
    cleaned = " ".join(value.strip().lower().replace("_", " ").split())
    for canonical, aliases in HEADER_ALIASES.items():
        if cleaned in aliases:
            return canonical
    return None


def _record_from_row(row: dict[str, str], path: Path) -> TenderRecord:
    title = row.get("title", "")
    url = row.get("url", "")
    number = row.get("number") or None
    if not title or not (url or number):
        raise ValueError(f"{path.name}: row requires title and url or number")
    source = _source_name(row.get("source") or path.stem)
    raw_text = row.get("raw_text") or " ".join(value for value in row.values() if value)
    return TenderRecord(
        title=title,
        url=url or f"import://{path.name}/{number}",
        source=source,
        tender_number=number,
        customer=row.get("customer") or None,
        region=row.get("region") or None,
        price=_parse_price(row.get("price")),
        deadline=_parse_datetime(row.get("deadline")),
        published_at=_parse_datetime(row.get("published_at")),
        status="Актуально",
        raw_text=raw_text,
        detail_status="imported",
        source_confidence=0.8,
    )


def _source_name(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", _transliterate(value).lower()).strip("-")
    return f"import-{slug or 'manual'}"


def _transliterate(value: str) -> str:
    mapping = {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "е": "e",
        "ё": "e",
        "ж": "zh",
        "з": "z",
        "и": "i",
        "й": "y",
        "к": "k",
        "л": "l",
        "м": "m",
        "н": "n",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "у": "u",
        "ф": "f",
        "х": "h",
        "ц": "c",
        "ч": "ch",
        "ш": "sh",
        "щ": "sch",
        "ъ": "",
        "ы": "y",
        "ь": "",
        "э": "e",
        "ю": "yu",
        "я": "ya",
    }
    return "".join(mapping.get(char, char) for char in value.lower())


def _parse_price(value: str | None) -> float | None:
    if not value:
        return None
    cleaned = re.sub(r"[^\d,.\-]", "", value)
    if not cleaned:
        return None
    if "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    text = value.strip()
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None
