from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tender_parser.documents import DocumentAnalyzer, SUPPORTED_SUFFIXES
from tender_parser.tender_case import CaseEconomics, CaseExpense, ProductOffer, TenderCase


TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
READY_FILL = PatternFill("solid", fgColor="C6EFCE")
REVIEW_FILL = PatternFill("solid", fgColor="FFEB9C")
STOP_FILL = PatternFill("solid", fgColor="FFC7CE")


def export_case_report(
    tender_case: TenderCase,
    economics: CaseEconomics,
    offers: list[ProductOffer],
    expenses: list[CaseExpense],
    case_dir: Path,
    output_path: Path,
) -> Path:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Решение"
    _build_summary(summary, tender_case, economics)
    _build_positions(workbook.create_sheet("Позиции"), economics)
    _build_offers(workbook.create_sheet("Предложения"), offers)
    _build_expenses(workbook.create_sheet("Расходы"), expenses)
    _build_entities(workbook.create_sheet("ООО и ИП"), economics)
    _build_bidding(workbook.create_sheet("Торги"), tender_case, economics)
    _build_list_sheet(workbook.create_sheet("Риски"), "Риск", economics.risks)
    _build_list_sheet(workbook.create_sheet("Вопросы"), "Вопрос / действие", economics.questions)
    _build_documents(workbook.create_sheet("Документы"), case_dir / "documents")
    _build_checklist(workbook.create_sheet("Комплект заявки"), tender_case, economics)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _build_summary(sheet, tender_case: TenderCase, economics: CaseEconomics) -> None:
    sheet["A1"] = "ТЕНДЕРНОЕ ДЕЛО — ПРЕДВАРИТЕЛЬНОЕ РЕШЕНИЕ"
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = TITLE_FILL
    sheet.merge_cells("A1:D1")
    rows = [
        ("Статус", _decision_label(economics.decision)),
        ("Причина", economics.decision_reason),
        ("Номер дела", tender_case.case_id),
        ("Закупка", tender_case.title),
        ("Номер закупки", tender_case.tender_number),
        ("Источник", tender_case.source_url),
        ("Закон / тип", tender_case.law),
        ("Заказчик", tender_case.customer),
        ("Регион", tender_case.region),
        ("Адрес поставки", tender_case.delivery_address),
        ("НМЦК", tender_case.nmck),
        ("Плановая цена предложения", tender_case.planned_bid),
        ("Закупка товара, с НДС", economics.procurement_gross),
        ("Дополнительные расходы", economics.expenses_gross),
        ("Цена для наценки 30%", economics.target_price),
        ("Рабочий порог 15%", economics.viable_price),
        ("Жесткий порог 12%", economics.hard_floor_price),
        ("Запас НМЦК/плановой цены к цели", economics.headroom_to_target),
        ("Срок оплаты, дней", tender_case.payment_days),
        ("Срок поставки, дней", tender_case.delivery_days),
        ("Примечание", tender_case.notes),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        sheet.cell(row_index, 1, label).font = Font(bold=True)
        sheet.cell(row_index, 1).fill = HEADER_FILL
        sheet.cell(row_index, 2, _excel_value(value))
        if isinstance(value, Decimal):
            sheet.cell(row_index, 2).number_format = '#,##0.00 "₽"'
    status_cell = sheet["B3"]
    status_cell.font = Font(bold=True)
    status_cell.fill = {"ready": READY_FILL, "manual_review": REVIEW_FILL}.get(economics.decision, STOP_FILL)
    if tender_case.source_url:
        sheet["B8"].hyperlink = tender_case.source_url
        sheet["B8"].style = "Hyperlink"
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 90
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 16
    sheet.freeze_panes = "A3"
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_positions(sheet, economics: CaseEconomics) -> None:
    headers = [
        "№",
        "Позиция ТЗ",
        "Количество",
        "Ед.",
        "Обязательные характеристики",
        "Поставщик",
        "Артикул",
        "Подобранный товар",
        "Соответствие",
        "Цена за ед. с НДС",
        "Стоимость",
        "Наличие",
        "Срок, дней",
        "Подтверждение",
        "Комментарий",
    ]
    _headers(sheet, headers)
    for row_index, line in enumerate(economics.selected_lines, start=2):
        offer = line.offer
        values = [
            line.item.line_id,
            line.item.name,
            float(line.item.quantity),
            line.item.unit,
            line.item.required_specs,
            offer.supplier if offer else "",
            offer.sku if offer else "",
            offer.product_name if offer else "НЕ НАЙДЕН",
            offer.compliance_status if offer else "not_found",
            float(offer.unit_cost_gross) if offer else None,
            float(line.total_cost_gross) if offer else None,
            offer.stock if offer else "",
            offer.lead_days if offer else None,
            offer.evidence if offer else "",
            offer.notes if offer else "Требуется поиск товара",
        ]
        sheet.append(values)
        sheet.cell(row_index, 10).number_format = '#,##0.00 "₽"'
        sheet.cell(row_index, 11).number_format = '#,##0.00 "₽"'
        if not offer:
            for cell in sheet[row_index]:
                cell.fill = STOP_FILL
        elif offer.compliance_status == "conditional":
            for cell in sheet[row_index]:
                cell.fill = REVIEW_FILL
    _finish_table(sheet, widths=[10, 34, 12, 10, 48, 24, 20, 42, 18, 18, 18, 18, 12, 48, 40])


def _build_offers(sheet, offers: list[ProductOffer]) -> None:
    headers = [
        "Позиция",
        "Выбрано",
        "Поставщик",
        "Артикул",
        "Товар",
        "Цена с НДС",
        "Статус соответствия",
        "Наличие",
        "Срок, дней",
        "Ссылка",
        "Подтверждение характеристик",
        "Комментарий",
    ]
    _headers(sheet, headers)
    for offer in offers:
        sheet.append(
            [
                offer.line_id,
                "Да" if offer.selected else "",
                offer.supplier,
                offer.sku,
                offer.product_name,
                float(offer.unit_cost_gross),
                offer.compliance_status,
                offer.stock,
                offer.lead_days,
                offer.source_url,
                offer.evidence,
                offer.notes,
            ]
        )
        row = sheet.max_row
        sheet.cell(row, 6).number_format = '#,##0.00 "₽"'
        if offer.source_url:
            sheet.cell(row, 10).hyperlink = offer.source_url
            sheet.cell(row, 10).style = "Hyperlink"
    _finish_table(sheet, widths=[10, 12, 24, 20, 42, 18, 20, 18, 12, 38, 48, 40])


def _build_expenses(sheet, expenses: list[CaseExpense]) -> None:
    _headers(sheet, ["Категория", "Описание", "Сумма", "НДС", "НДС к вычету", "Подтверждено", "Комментарий"])
    for expense in expenses:
        sheet.append(
            [
                expense.category,
                expense.description,
                float(expense.amount_gross),
                float(expense.vat_rate),
                "Да" if expense.vat_reclaimable else "Нет",
                "Да" if expense.confirmed else "Нет",
                expense.notes,
            ]
        )
        sheet.cell(sheet.max_row, 3).number_format = '#,##0.00 "₽"'
        sheet.cell(sheet.max_row, 4).number_format = "0%"
        if not expense.confirmed:
            for cell in sheet[sheet.max_row]:
                cell.fill = REVIEW_FILL
    _finish_table(sheet, widths=[18, 34, 18, 12, 16, 16, 48])


def _build_entities(sheet, economics: CaseEconomics) -> None:
    _headers(
        sheet,
        [
            "Сценарий",
            "Цена для заказчика",
            "Выручка без НДС",
            "Закупка",
            "Расходы",
            "Прибыль до налога",
            "Оценочный налог",
            "Минимальный налог 1% (справочно)",
            "Прибыль после оценочного налога",
            "Доходность затрат",
            "Примечание",
        ],
    )
    for scenario in economics.entity_scenarios:
        sheet.append(
            [
                scenario.entity,
                float(scenario.sale_price_gross),
                float(scenario.revenue_net),
                float(scenario.procurement_cost),
                float(scenario.expense_cost),
                float(scenario.profit_before_income_tax),
                float(scenario.estimated_tax) if scenario.estimated_tax is not None else None,
                float(scenario.minimum_tax_reference) if scenario.minimum_tax_reference is not None else None,
                float(scenario.profit_after_estimated_tax) if scenario.profit_after_estimated_tax is not None else None,
                float(scenario.profit_rate_on_cost) if scenario.profit_rate_on_cost is not None else None,
                scenario.note,
            ]
        )
        for column in range(2, 10):
            sheet.cell(sheet.max_row, column).number_format = '#,##0.00 "₽"'
        sheet.cell(sheet.max_row, 10).number_format = "0.00%"
    _finish_table(sheet, widths=[30, 20, 20, 20, 20, 20, 18, 24, 24, 18, 76])


def _build_bidding(sheet, tender_case: TenderCase, economics: CaseEconomics) -> None:
    _headers(sheet, ["Уровень", "Минимальная цена", "Снижение от НМЦК, ₽", "Снижение от НМЦК, %", "Решение"])
    levels = [
        ("Цель 30%", economics.target_price, economics.target_discount_from_nmck, "Рабочая цель торгов"),
        ("Рабочий порог 15%", economics.viable_price, economics.viable_discount_from_nmck, "Ниже — только ручное решение"),
        ("Жесткий порог 12%", economics.hard_floor_price, economics.hard_floor_discount_from_nmck, "Ниже не торговаться"),
    ]
    for label, threshold, discount, decision in levels:
        reduction_amount = tender_case.nmck - threshold if tender_case.nmck is not None else None
        sheet.append(
            [
                label,
                float(threshold),
                float(reduction_amount) if reduction_amount is not None else None,
                float(discount) if discount is not None else None,
                decision,
            ]
        )
        row = sheet.max_row
        sheet.cell(row, 2).number_format = '#,##0.00 "₽"'
        sheet.cell(row, 3).number_format = '#,##0.00 "₽"'
        sheet.cell(row, 4).number_format = "0.00%"
        if label.startswith("Жесткий"):
            for cell in sheet[row]:
                cell.fill = STOP_FILL
        elif label.startswith("Рабочий"):
            for cell in sheet[row]:
                cell.fill = REVIEW_FILL
        else:
            for cell in sheet[row]:
                cell.fill = READY_FILL
    sheet.append([])
    sheet.append(["Текущая оценочная цена", float(economics.assessment_price) if economics.assessment_price else None, "", "", economics.decision_reason])
    sheet.cell(sheet.max_row, 2).number_format = '#,##0.00 "₽"'
    _finish_table(sheet, widths=[26, 22, 24, 24, 56])


def _build_list_sheet(sheet, header: str, values: list[str]) -> None:
    _headers(sheet, ["№", header, "Статус/комментарий владельца"])
    if not values:
        sheet.append([1, "Нет автоматически выявленных пунктов", ""])
    else:
        for index, value in enumerate(values, start=1):
            sheet.append([index, value, ""])
    _finish_table(sheet, widths=[8, 90, 48])


def _build_documents(sheet, documents_dir: Path) -> None:
    _headers(sheet, ["Файл", "Тип", "Размер, КБ", "Извлечен текст", "Комментарий"])
    evidence = DocumentAnalyzer(documents_dir).analyze()
    paths = []
    if documents_dir.exists():
        paths = [path for path in sorted(documents_dir.rglob("*")) if path.is_file()]
    for path in paths:
        supported = path.suffix.lower() in SUPPORTED_SUFFIXES
        sheet.append(
            [
                str(path.relative_to(documents_dir)),
                path.suffix.lower(),
                round(path.stat().st_size / 1024, 1),
                "Да" if supported else "Нет",
                "Нужен OCR" if path.suffix.lower() == ".pdf" and not evidence.searchable_text else "",
            ]
        )
    if not paths:
        sheet.append(["", "", "", "Нет", "Документы закупки еще не загружены"])
    sheet.append([])
    sheet.append(["Сводка распознавания", evidence.summary])
    _finish_table(sheet, widths=[58, 12, 14, 18, 48])


def _build_checklist(sheet, tender_case: TenderCase, economics: CaseEconomics) -> None:
    _headers(sheet, ["Блок", "Документ / проверка", "Статус", "Комментарий"])
    rows = [
        ("Закупка", "Извещение и действующая редакция документов", "Проверить", "Учитывать все изменения заказчика"),
        ("Закупка", "Техническое задание / спецификация", "Проверить", "Разложить на обязательные характеристики"),
        ("Закупка", "Проект контракта", "Проверить", "Сроки, приемка, штрафы, гарантия, оплата"),
        ("Участник", "Карточка предприятия", "Подготовить", "Выбрать ООО или ИП"),
        ("Участник", "Декларации и согласия", "Подготовить", "По требованиям конкретной процедуры"),
        ("Товар", "Таблица соответствия", "Готово" if economics.selected_lines else "Не готово", "Только с доказательствами характеристик"),
        ("Товар", "Сертификаты / декларации", "Проверить", "Если требуются документацией"),
        ("Цена", "Калькуляция", "Готово" if economics.procurement_gross > 0 else "Не готово", economics.decision_reason),
        ("Коммерческое", "Коммерческое предложение по шаблону", "После согласования", "Печатает и подписывает владелец"),
        ("Подача", "Финальная ручная проверка", "Обязательно", "Автоматическая подача и ЭЦП отключены"),
    ]
    if tender_case.law.lower().startswith("223"):
        rows.insert(3, ("Право", "Положение о закупке заказчика", "Проверить", "Обязательно для 223-ФЗ"))
    for row in rows:
        sheet.append(row)
    _finish_table(sheet, widths=[18, 54, 20, 68])


def _headers(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _finish_table(sheet, widths: list[int]) -> None:
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 2:
        sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _decision_label(value: str) -> str:
    return {
        "ready": "МОЖНО РАССМАТРИВАТЬ УЧАСТИЕ",
        "manual_review": "НУЖНО РЕШЕНИЕ ВЛАДЕЛЬЦА",
        "stop": "ЭКОНОМИЧЕСКИЙ СТОП",
        "blocked": "АНАЛИЗ НЕ ЗАВЕРШЕН",
    }.get(value, value)


def _excel_value(value):
    if isinstance(value, Decimal):
        return float(value)
    return value
