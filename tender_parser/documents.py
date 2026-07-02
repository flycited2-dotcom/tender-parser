from __future__ import annotations

import csv
import json
from zipfile import ZipFile
from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader

from tender_parser.config import BROAD_SEARCH_TERMS, CATEGORY_KEYWORDS, REGION_TERMS, STOP_TERMS
from tender_parser.text import normalize_text, word_term_matches


SUPPORTED_SUFFIXES = {".csv", ".docx", ".html", ".htm", ".json", ".pdf", ".txt", ".xlsx", ".xml"}


@dataclass(frozen=True)
class DocumentEvidence:
    matched_terms: list[str] = field(default_factory=list)
    regions: list[str] = field(default_factory=list)
    stop_terms: list[str] = field(default_factory=list)
    summary: str = ""
    searchable_text: str = ""


class DocumentAnalyzer:
    def __init__(self, documents_dir: Path) -> None:
        self.documents_dir = documents_dir

    def analyze(self) -> DocumentEvidence:
        if not self.documents_dir.exists():
            return DocumentEvidence()

        terms: list[str] = []
        regions: list[str] = []
        stop_terms: list[str] = []
        summaries: list[str] = []
        all_text: list[str] = []

        for path in sorted(self.documents_dir.rglob("*")):
            if not path.is_file() or path.suffix.lower() not in SUPPORTED_SUFFIXES:
                continue
            text = _read_document_text(path)
            if not text.strip():
                continue
            all_text.append(text)
            file_terms = _matching_terms(text, _target_terms())
            file_regions = _matching_terms(text, REGION_TERMS)
            file_stop_terms = _matching_terms(text, STOP_TERMS)
            terms.extend(file_terms)
            regions.extend(file_regions)
            stop_terms.extend(file_stop_terms)
            if file_terms or file_regions or file_stop_terms:
                parts = []
                if file_regions:
                    parts.append(f"regions={', '.join(file_regions)}")
                if file_terms:
                    parts.append(f"terms={', '.join(file_terms)}")
                if file_stop_terms:
                    parts.append(f"stop={', '.join(file_stop_terms)}")
                summaries.append(f"{path.name}: {'; '.join(parts)}")

        return DocumentEvidence(
            matched_terms=_unique(terms),
            regions=_unique(regions),
            stop_terms=_unique(stop_terms),
            summary=" | ".join(summaries),
            searchable_text=normalize_text(" ".join(all_text)),
        )


def _target_terms() -> list[str]:
    terms: list[str] = []
    for category_terms in CATEGORY_KEYWORDS.values():
        terms.extend(category_terms)
    terms.extend(BROAD_SEARCH_TERMS)
    return _unique(normalize_text(term) for term in terms)


def _matching_terms(text: str, terms: list[str]) -> list[str]:
    normalized_text = normalize_text(text)
    return [normalize_text(term) for term in terms if _term_matches(normalized_text, normalize_text(term))]


def _term_matches(text: str, term: str) -> bool:
    if not term:
        return False
    if " " not in term:
        return word_term_matches(text, term)
    if term in text:
        return True
    return _phrase_stems_match(text, term)


def _phrase_stems_match(text: str, term: str) -> bool:
    stems = [_stem(word) for word in term.split() if len(_stem(word)) >= 4]
    if not stems:
        return False
    position = 0
    for stem in stems:
        found_at = text.find(stem, position)
        if found_at < 0:
            return False
        position = found_at + len(stem)
    return True


def _stem(word: str) -> str:
    return word[: max(4, min(7, len(word) - 2))]


def _read_document_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_text(path)
    if suffix == ".docx":
        return _read_docx_text(path)
    if suffix in {".html", ".htm"}:
        return _strip_html(_read_text(path))
    if suffix == ".json":
        return _read_json_text(path)
    if suffix == ".pdf":
        return _read_pdf_text(path)
    if suffix == ".xlsx":
        return _read_xlsx_text(path)
    if suffix == ".xml":
        return " ".join(element.text or "" for element in ElementTree.fromstring(_read_text(path)).iter())
    return _read_text(path)


def _read_csv_text(path: Path) -> str:
    text = _read_text(path)
    delimiter = ";" if text.count(";") >= text.count(",") else ","
    rows = csv.reader(text.splitlines(), delimiter=delimiter)
    return " ".join(" ".join(cell for cell in row if cell) for row in rows)


def _read_json_text(path: Path) -> str:
    payload = json.loads(_read_text(path))
    return _flatten_json(payload)


def _read_xlsx_text(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        values: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                values.extend(str(cell) for cell in row if cell is not None)
        return " ".join(values)
    finally:
        workbook.close()


def _read_docx_text(path: Path) -> str:
    with ZipFile(path) as archive:
        xml_text = archive.read("word/document.xml").decode("utf-8", errors="ignore")
    root = ElementTree.fromstring(xml_text)
    return " ".join(element.text or "" for element in root.iter())


def _read_pdf_text(path: Path) -> str:
    reader = PdfReader(path)
    values: list[str] = []
    for page in reader.pages:
        values.append(page.extract_text() or "")
        for annotation in page.annotations or []:
            value = annotation.get_object().get("/Contents")
            if value:
                values.append(str(value))
    return " ".join(values)


def _flatten_json(value: object) -> str:
    if isinstance(value, dict):
        return " ".join(_flatten_json(item) for item in value.values())
    if isinstance(value, list):
        return " ".join(_flatten_json(item) for item in value)
    return str(value) if value is not None else ""


def _read_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def _strip_html(value: str) -> str:
    parser = _TextExtractor()
    parser.feed(value)
    return parser.text


class _TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    @property
    def text(self) -> str:
        return " ".join(self.parts)

    def handle_data(self, data: str) -> None:
        if data.strip():
            self.parts.append(data.strip())


def _unique(values: object) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value)
        if text and text not in seen:
            seen.add(text)
            result.append(text)
    return result
