from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from tender_parser.models import TenderRecord
from tender_parser.text import normalize_text, stem


INDEX_SCHEMA_VERSION = 1
MAX_HEADER_SCAN_ROWS = 20
QUERY_STOP_WORDS = {
    "44",
    "223",
    "а",
    "в",
    "для",
    "и",
    "из",
    "к",
    "комплект",
    "комплекта",
    "комплектов",
    "на",
    "по",
    "поставка",
    "поставки",
    "поставку",
    "приобретение",
    "республика",
    "с",
    "товар",
    "товара",
    "товаров",
    "услуга",
    "услуги",
    "услуг",
    "фз",
}


@dataclass(frozen=True)
class SupplierDefinition:
    supplier_id: str
    name: str
    email: str = ""
    email_senders: tuple[str, ...] = ()
    website: str = ""
    enabled: bool = True
    file_globs: tuple[str, ...] = ()
    tender_categories: tuple[str, ...] = ()
    priority: int = 100


@dataclass(frozen=True)
class SupplierProduct:
    supplier_id: str
    supplier_name: str
    article: str
    name: str
    category: str
    dealer_price: float | None = None
    retail_price: float | None = None
    status: str = ""
    source_file: str = ""
    source_sheet: str = ""

    @property
    def preferred_price(self) -> float | None:
        return self.dealer_price if self.dealer_price is not None else self.retail_price

    @property
    def price_type(self) -> str:
        return "дилерская" if self.dealer_price is not None else "розничная"


@dataclass(frozen=True)
class SupplierMatch:
    product: SupplierProduct
    score: float
    matched_tokens: tuple[str, ...]


@dataclass(frozen=True)
class SupplierIndexStatus:
    status: str
    product_count: int = 0
    supplier_count: int = 0
    file_count: int = 0
    detail: str = ""


class SupplierCatalog:
    def __init__(self, catalog_dir: Path) -> None:
        self.catalog_dir = catalog_dir.resolve()
        self.private_dir = self.catalog_dir / "private"
        self.manifest_path = self.catalog_dir / "suppliers.json"
        self.index_path = self.private_dir / "catalog_index.json"
        self.products: list[SupplierProduct] = []
        self.definitions: dict[str, SupplierDefinition] = {}
        self.last_status = SupplierIndexStatus(status="not_loaded")

    def refresh(self, *, force: bool = False) -> SupplierIndexStatus:
        try:
            definitions = self._load_definitions()
        except (OSError, ValueError, TypeError, KeyError) as exc:
            self.products = []
            self.definitions = {}
            self.last_status = SupplierIndexStatus(
                status="error", detail=f"не удалось прочитать suppliers.json: {exc.__class__.__name__}"
            )
            return self.last_status
        self.definitions = {item.supplier_id: item for item in definitions if item.enabled}
        if not self.definitions:
            self.products = []
            self.last_status = SupplierIndexStatus(status="disabled", detail="нет активных поставщиков")
            return self.last_status

        sources = self._source_files(self.definitions.values())
        fingerprints = {
            self._relative(path): {"size": path.stat().st_size, "mtime_ns": path.stat().st_mtime_ns}
            for _, path in sources
        }
        if not force and self._load_current_index(fingerprints):
            return self.last_status

        products: dict[tuple[str, str], SupplierProduct] = {}
        errors: list[str] = []
        for supplier, path in sources:
            try:
                for product in _read_supplier_workbook(path, supplier):
                    key = (product.supplier_id, normalize_text(product.article or product.name))
                    products[key] = _merge_product(products.get(key), product)
            except (OSError, ValueError, TypeError) as exc:
                errors.append(f"{path.name}: {exc.__class__.__name__}")

        self.products = sorted(
            products.values(),
            key=lambda item: (item.supplier_name.casefold(), item.category.casefold(), item.name.casefold()),
        )
        self.private_dir.mkdir(parents=True, exist_ok=True)
        payload = {
            "schema_version": INDEX_SCHEMA_VERSION,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "fingerprints": fingerprints,
            "products": [asdict(item) for item in self.products],
        }
        temporary = self.index_path.with_suffix(".tmp")
        temporary.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        temporary.replace(self.index_path)
        status = "partial" if errors else ("ok" if sources else "empty")
        detail = "; ".join(errors) if errors else ""
        self.last_status = SupplierIndexStatus(
            status=status,
            product_count=len(self.products),
            supplier_count=len(self.definitions),
            file_count=len(sources),
            detail=detail,
        )
        return self.last_status

    def search(
        self,
        query: str,
        *,
        limit: int = 10,
        supplier_id: str | None = None,
        tender_category: str | None = None,
    ) -> list[SupplierMatch]:
        if self.last_status.status == "not_loaded":
            self.refresh()
        query_normalized = normalize_text(query)
        query_tokens = _search_tokens(query_normalized)
        if not query_tokens:
            return []
        query_stems = {stem(token) for token in query_tokens}
        results: list[SupplierMatch] = []
        for product in self.products:
            if supplier_id and product.supplier_id != supplier_id:
                continue
            definition = self.definitions.get(product.supplier_id)
            if (
                tender_category
                and definition
                and definition.tender_categories
                and tender_category not in definition.tender_categories
            ):
                continue
            score, matched = _score_product(product, query_normalized, query_tokens, query_stems)
            if score < 25:
                continue
            results.append(SupplierMatch(product=product, score=score, matched_tokens=matched))
        results.sort(
            key=lambda item: (
                -item.score,
                self.definitions.get(item.product.supplier_id, SupplierDefinition("", "")).priority,
                item.product.preferred_price is None,
                item.product.preferred_price or 0,
                item.product.name.casefold(),
            )
        )
        return results[: max(1, min(limit, 50))]

    def match_tenders(
        self, tenders: Iterable[TenderRecord], *, limit_per_tender: int = 5
    ) -> list[dict[str, object]]:
        if self.last_status.status == "not_loaded":
            self.refresh()
        matches: list[dict[str, object]] = []
        for tender in tenders:
            candidates = self.search(
                " ".join([tender.title, *tender.matched_terms]),
                limit=limit_per_tender,
                tender_category=tender.category,
            )
            if not candidates:
                continue
            matches.append(
                {
                    "tender_key": tender.unique_key,
                    "tender_number": tender.official_number or tender.platform_number or tender.tender_number,
                    "tender_title": tender.title,
                    "tender_url": tender.official_url or tender.platform_url or tender.url,
                    "category": tender.category,
                    "candidates": [_match_dict(item) for item in candidates],
                }
            )
        return matches

    def _load_definitions(self) -> list[SupplierDefinition]:
        payloads = [json.loads(self.manifest_path.read_text(encoding="utf-8"))]
        auto_manifest = self.private_dir / "suppliers_auto.json"
        if auto_manifest.is_file():
            payloads.append(json.loads(auto_manifest.read_text(encoding="utf-8")))
        result: dict[str, SupplierDefinition] = {}
        for payload in reversed(payloads):
            if payload.get("schema_version") != 1:
                raise ValueError("unsupported schema")
            for raw in payload.get("suppliers", []):
                supplier_id = str(raw.get("id", "")).strip().casefold()
                name = str(raw.get("name", "")).strip()
                if not supplier_id or not name:
                    continue
                email = str(raw.get("email", "")).strip()
                senders = raw.get("email_senders", [email] if email else [])
                definition = SupplierDefinition(
                    supplier_id=supplier_id,
                    name=name,
                    email=email,
                    email_senders=tuple(
                        str(value).strip()
                        for value in senders
                        if str(value).strip()
                    ),
                    website=str(raw.get("website", "")).strip(),
                    enabled=bool(raw.get("enabled", True)),
                    file_globs=tuple(str(value) for value in raw.get("file_globs", [])),
                    tender_categories=tuple(
                        str(value) for value in raw.get("tender_categories", [])
                    ),
                    priority=max(1, int(raw.get("priority", 100))),
                )
                result[definition.supplier_id] = definition
        return list(result.values())

    def _source_files(
        self, definitions: Iterable[SupplierDefinition]
    ) -> list[tuple[SupplierDefinition, Path]]:
        result: list[tuple[SupplierDefinition, Path]] = []
        private_root = self.private_dir.resolve()
        for supplier in definitions:
            for pattern in supplier.file_globs:
                for path in self.catalog_dir.glob(pattern):
                    resolved = path.resolve()
                    if (
                        resolved.is_file()
                        and resolved.suffix.casefold() in {".xlsx", ".xlsm"}
                        and not resolved.name.startswith("~$")
                        and resolved.is_relative_to(private_root)
                    ):
                        result.append((supplier, resolved))
        return sorted(set(result), key=lambda item: (item[0].supplier_id, str(item[1])))

    def _load_current_index(self, fingerprints: dict[str, dict[str, int]]) -> bool:
        try:
            payload = json.loads(self.index_path.read_text(encoding="utf-8"))
            if payload.get("schema_version") != INDEX_SCHEMA_VERSION:
                return False
            if payload.get("fingerprints") != fingerprints:
                return False
            self.products = [SupplierProduct(**item) for item in payload.get("products", [])]
        except (OSError, ValueError, TypeError):
            return False
        self.last_status = SupplierIndexStatus(
            status="ok" if self.products else "empty",
            product_count=len(self.products),
            supplier_count=len(self.definitions),
            file_count=len(fingerprints),
        )
        return True

    def _relative(self, path: Path) -> str:
        return path.relative_to(self.catalog_dir).as_posix()


def export_supplier_matches(
    catalog: SupplierCatalog,
    tenders: Iterable[TenderRecord],
    output_path: Path,
) -> tuple[Path, int]:
    matches = catalog.match_tenders(tenders)
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "catalog_status": asdict(catalog.last_status),
        "matched_tenders": len(matches),
        "items": matches,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path, len(matches)


def format_supplier_matches(query: str, matches: list[SupplierMatch]) -> str:
    if not matches:
        return f"По запросу «{query}» в локальных прайсах ничего не найдено."
    lines = [f"Прайсы поставщиков: {query}", f"Найдено вариантов: {len(matches)}"]
    for index, match in enumerate(matches, 1):
        product = match.product
        price = product.preferred_price
        price_text = (
            f"{price:,.2f} ₽".replace(",", " ") if price is not None else "цена не указана"
        )
        lines.extend(
            [
                "",
                f"{index}. {product.supplier_name} — {product.name}",
                f"Артикул: {product.article or '—'} | {product.price_type}: {price_text}",
                f"Раздел: {product.category}",
            ]
        )
    return "\n".join(lines)


def _read_supplier_workbook(
    path: Path, supplier: SupplierDefinition
) -> Iterable[SupplierProduct]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = sorted(workbook.worksheets, key=lambda sheet: sheet.title.casefold() == "база")
        for sheet in sheets:
            columns = _find_columns(sheet)
            if columns is None:
                continue
            for row in sheet.iter_rows(min_row=columns["header_row"] + 1, values_only=True):
                article = _cell_text(row, columns["article"])
                name = _cell_text(row, columns["name"])
                if not article or not name:
                    continue
                dealer_price = _cell_price(row, columns.get("dealer_price"))
                retail_price = _cell_price(row, columns.get("retail_price"))
                generic_price = _cell_price(row, columns.get("generic_price"))
                if dealer_price is None and retail_price is None:
                    retail_price = generic_price
                if dealer_price is not None and dealer_price <= 0:
                    dealer_price = None
                if retail_price is not None and retail_price <= 0:
                    retail_price = None
                yield SupplierProduct(
                    supplier_id=supplier.supplier_id,
                    supplier_name=supplier.name,
                    article=article,
                    name=name,
                    category="" if normalize_text(sheet.title) == "база" else sheet.title,
                    dealer_price=dealer_price,
                    retail_price=retail_price,
                    status=_cell_text(row, columns.get("status")),
                    source_file=path.name,
                    source_sheet=sheet.title,
                )
    finally:
        workbook.close()


def _find_columns(sheet: object) -> dict[str, int] | None:
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(MAX_HEADER_SCAN_ROWS, sheet.max_row), values_only=True),
        1,
    ):
        headers = [_header(value) for value in row]
        article = _first_index(headers, {"артикул", "sku", "код товара"})
        name = _first_index(headers, {"наименование", "название", "товар", "продукция"})
        if article is None or name is None:
            continue
        result: dict[str, int] = {"header_row": row_number, "article": article, "name": name}
        for index, header in enumerate(headers):
            if (
                "дилер" in header
                or "цена закупки" in header
                or "закупочная цена" in header
                or ("оптов" in header and "цен" in header)
            ):
                result["dealer_price"] = index
            elif "рознич" in header and "цен" in header:
                result["retail_price"] = index
            elif header in {"цена", "стоимость"}:
                result["generic_price"] = index
            elif header.startswith("статус"):
                result["status"] = index
        return result
    return None


def _merge_product(
    current: SupplierProduct | None, incoming: SupplierProduct
) -> SupplierProduct:
    if current is None:
        return incoming
    preferred = incoming if incoming.dealer_price is not None else current
    alternate = current if preferred is incoming else incoming
    return replace(
        preferred,
        category=preferred.category or alternate.category,
        dealer_price=preferred.dealer_price or alternate.dealer_price,
        retail_price=preferred.retail_price or alternate.retail_price,
        status=preferred.status or alternate.status,
    )


def _score_product(
    product: SupplierProduct,
    query_normalized: str,
    query_tokens: tuple[str, ...],
    query_stems: set[str],
) -> tuple[float, tuple[str, ...]]:
    name = normalize_text(product.name)
    article = normalize_text(product.article)
    category = normalize_text(product.category)
    product_tokens = _search_tokens(" ".join((name, article)), keep_stop_words=True)
    product_stems = {stem(token) for token in product_tokens}
    matched_stems = query_stems & product_stems
    if not matched_stems:
        return 0.0, ()
    model_tokens = {token for token in query_tokens if any(char.isdigit() for char in token)}
    if model_tokens and not model_tokens.intersection(product_tokens):
        return 0.0, ()
    coverage = len(matched_stems) / len(query_stems)
    score = coverage * 70 + len(matched_stems) * 2
    if len(query_normalized) >= 4 and query_normalized in name:
        score += 35
    if article and (query_normalized == article or article in query_tokens):
        score += 120
    score += len(model_tokens & set(product_tokens)) * 25
    if product_tokens and stem(product_tokens[0]) in query_stems:
        score += 20
    category_stems = {stem(token) for token in _search_tokens(category, keep_stop_words=True)}
    if query_stems & category_stems:
        score += 8
    matched_tokens = tuple(token for token in query_tokens if stem(token) in matched_stems)
    return round(score, 2), matched_tokens


def _search_tokens(value: str, *, keep_stop_words: bool = False) -> tuple[str, ...]:
    tokens = re.findall(r"[a-zа-я0-9]+(?:[-/.][a-zа-я0-9]+)*", normalize_text(value))
    result = []
    for token in tokens:
        if len(token) < 2:
            continue
        if not keep_stop_words and token in QUERY_STOP_WORDS:
            continue
        if token not in result:
            result.append(token)
    return tuple(result)


def _header(value: object) -> str:
    return re.sub(r"[^a-zа-я0-9]+", " ", normalize_text(str(value) if value is not None else "")).strip()


def _first_index(values: list[str], options: set[str]) -> int | None:
    return next((index for index, value in enumerate(values) if value in options), None)


def _cell_text(row: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _cell_price(row: tuple[object, ...], index: int | None) -> float | None:
    if index is None or index >= len(row):
        return None
    value = row[index]
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    cleaned = re.sub(r"[^0-9,.-]+", "", str(value)).replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _match_dict(match: SupplierMatch) -> dict[str, object]:
    product = match.product
    return {
        "supplier_id": product.supplier_id,
        "supplier_name": product.supplier_name,
        "article": product.article,
        "name": product.name,
        "category": product.category,
        "dealer_price": product.dealer_price,
        "retail_price": product.retail_price,
        "preferred_price": product.preferred_price,
        "price_type": product.price_type,
        "score": match.score,
        "matched_tokens": list(match.matched_tokens),
        "source_file": product.source_file,
        "source_sheet": product.source_sheet,
    }
