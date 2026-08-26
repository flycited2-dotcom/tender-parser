from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from tender_parser import config


class SearchConfigError(ValueError):
    pass


@dataclass(frozen=True)
class SearchProfile:
    category_keywords: dict[str, list[str]]
    search_terms: list[str]
    stop_terms: list[str]
    regions: list[str]
    min_price_rub: int


@dataclass(frozen=True)
class SearchConfigStatus:
    status: str
    path: Path
    detail: str = ""


DEFAULT_SEARCH_PROFILE = SearchProfile(
    category_keywords={
        category: list(terms) for category, terms in config.CATEGORY_KEYWORDS.items()
    },
    search_terms=list(config.SEARCH_QUERY_TERMS),
    stop_terms=list(config.STOP_TERMS),
    regions=list(config.SEARCH_REGION_TERMS),
    min_price_rub=config.MIN_PRICE_RUB,
)


def load_search_profile(path: Path) -> SearchProfile:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise SearchConfigError(f"не удалось открыть Excel: {exc.__class__.__name__}") from exc

    try:
        categories: dict[str, list[str]] = {}
        for row in _active_rows(workbook["Категории"], ["Категория", "Ключевое слово"]):
            category, term = row
            categories.setdefault(category, [])
            if term not in categories[category]:
                categories[category].append(term)

        search_terms = _single_value_rows(workbook["Поисковые запросы"], "Ключевое слово")
        stop_terms = _single_value_rows(workbook["Исключения"], "Стоп-слово или фраза")
        regions = _single_value_rows(workbook["Регионы"], "Регион")
        min_price = _read_min_price(workbook["Параметры"])
    except KeyError as exc:
        raise SearchConfigError(f"отсутствует лист или столбец: {exc.args[0]}") from exc
    finally:
        workbook.close()

    if not categories:
        raise SearchConfigError("нет активных категорий")
    if not search_terms:
        raise SearchConfigError("нет активных поисковых запросов")
    if not regions:
        raise SearchConfigError("нет активных регионов")
    return SearchProfile(
        category_keywords=categories,
        search_terms=search_terms,
        stop_terms=stop_terms,
        regions=regions,
        min_price_rub=min_price,
    )


def apply_search_profile(profile: SearchProfile) -> None:
    config.CATEGORY_KEYWORDS.clear()
    config.CATEGORY_KEYWORDS.update(
        {category: list(terms) for category, terms in profile.category_keywords.items()}
    )
    config.STOP_TERMS[:] = profile.stop_terms
    config.SEARCH_QUERY_TERMS[:] = profile.search_terms
    config.SEARCH_REGION_TERMS[:] = profile.regions
    regional = [f"{term} {region}" for term in profile.search_terms for region in profile.regions]
    config.REGIONAL_SEARCH_QUERIES[:] = regional
    # Эти имена исторически могли перестать быть алиасами после импорта/рефакторинга.
    config.ROSTENDER_SEARCH_QUERIES[:] = [
        *config.CUSTOMER_DISCOVERY_REGION_QUERIES,
        *regional,
    ]
    config.ETP_GPB_SEARCH_QUERIES[:] = [
        *config.CUSTOMER_DISCOVERY_REGION_QUERIES,
        *regional,
    ]
    config.EIS_SEARCH_QUERIES[:] = regional
    config.B2B_SEARCH_QUERIES[:] = [
        *config.CUSTOMER_DISCOVERY_REGION_QUERIES,
        *profile.search_terms,
    ]
    config.MIN_PRICE_RUB = profile.min_price_rub


def load_and_apply_search_config(path: Path) -> SearchConfigStatus:
    if not path.exists():
        apply_search_profile(DEFAULT_SEARCH_PROFILE)
        return SearchConfigStatus(status="missing", path=path, detail="используется встроенный словарь")
    try:
        profile = load_search_profile(path)
    except SearchConfigError as exc:
        apply_search_profile(DEFAULT_SEARCH_PROFILE)
        return SearchConfigStatus(status="error", path=path, detail=str(exc))
    apply_search_profile(profile)
    return SearchConfigStatus(
        status="loaded",
        path=path,
        detail=(
            f"категорий {len(profile.category_keywords)}, запросов {len(profile.search_terms)}, "
            f"исключений {len(profile.stop_terms)}, регионов {len(profile.regions)}"
        ),
    )


def _active_rows(sheet: object, required_headers: list[str]) -> list[list[str]]:
    headers = {
        str(cell.value).strip(): index
        for index, cell in enumerate(sheet[3])  # type: ignore[index]
        if cell.value is not None
    }
    active_index = headers.get("Активно")
    if active_index is None:
        raise KeyError("Активно")
    indexes: list[int] = []
    for header in required_headers:
        if header not in headers:
            raise KeyError(header)
        indexes.append(headers[header])

    result: list[list[str]] = []
    for values in sheet.iter_rows(min_row=4, values_only=True):  # type: ignore[attr-defined]
        if not _is_active(values[active_index] if active_index < len(values) else None):
            continue
        row = [_clean(values[index] if index < len(values) else None) for index in indexes]
        if all(row):
            result.append(row)
    return result


def _single_value_rows(sheet: object, header: str) -> list[str]:
    return _dedupe(row[0] for row in _active_rows(sheet, [header]))


def _read_min_price(sheet: object) -> int:
    for values in sheet.iter_rows(min_row=4, values_only=True):  # type: ignore[attr-defined]
        if _clean(values[0] if values else None).lower() != "минимальная сумма, руб.":
            continue
        value = values[1] if len(values) > 1 else None
        try:
            parsed = int(float(value))
        except (TypeError, ValueError) as exc:
            raise SearchConfigError("минимальная сумма должна быть числом") from exc
        if parsed < 0:
            raise SearchConfigError("минимальная сумма не может быть отрицательной")
        return parsed
    raise SearchConfigError("в листе «Параметры» нет минимальной суммы")


def _is_active(value: object) -> bool:
    return _clean(value).lower() in {"да", "yes", "true", "1", "on"}


def _clean(value: object) -> str:
    return " ".join(str(value).split()) if value is not None else ""


def _dedupe(values: object) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:  # type: ignore[union-attr]
        normalized = str(value).casefold()
        if normalized in seen:
            continue
        seen.add(normalized)
        result.append(str(value))
    return result
